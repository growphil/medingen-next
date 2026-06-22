from flask import Flask
from flask import request
from flask import jsonify
from flask import make_response
from flask import Response
from flask import redirect
from flask import send_file
from flask import render_template_string

import math
from datetime import datetime, timezone
from flask_cors import CORS

from datetime import datetime
from datetime import date
from datetime import timedelta
from datetime import timezone
from datetime import time as dt_time

from decimal import Decimal
from decimal import ROUND_HALF_UP

from random import randint
from zoneinfo import ZoneInfo

from functools import wraps
from xml.sax.saxutils import escape
from io import BytesIO

import re
import io
import tempfile
import subprocess
import shutil
import json
import math
import time
import zipfile
import random
import traceback
import hmac
import hashlib

import pymysql
import jwt
import boto3
import razorpay
import requests
import pytz

from openpyxl import Workbook
from openpyxl import load_workbook

from botocore.client import Config
from botocore.exceptions import NoCredentialsError
from botocore.exceptions import PartialCredentialsError
from boto3.session import Session

CORS_HEADERS = ['Content-Type', 'Authorization']
app = Flask(__name__, static_folder="static")
CORS(app)  


app.config['MYSQL_HOST'] = 'medingen-mysql-db-new.czes2c8i214u.ap-south-1.rds.amazonaws.com'
app.config['MYSQL_USER'] = 'admin'
app.config['MYSQL_PASSWORD'] = 'Medingen#2025!'
# app.config['MYSQL_DB'] = 'local_db'
app.config['MYSQL_DB'] = 'medingen_db'


try:
    session = Session(profile_name='kyra')
    s3_client = session.client('s3', region_name='ap-south-1', config=Config(signature_version='s3v4'))
    print("Using 'kyra' AWS profile for S3 operations")
except Exception as e:
    s3_client = boto3.client('s3', region_name='ap-south-1', config=Config(signature_version='s3v4'))
    print("Using default AWS credentials for S3 operations")

BUCKET_NAME = 'medingen-store-new'

app.config['SECRET_KEY'] = 'skandagn-secret'

redirect_map = None

key_id = "rzp_live_gD2p7TIiAYtioQ" # rzp_test_xebDqhKAMJGlGu
key_secret = "M8g5TpPWAAAtk6MZatBfJOVp" # wq8IlTphjxbaU2U6m6hlhjLk
client = razorpay.Client(auth=(key_id, key_secret))
RAZORPAY_SECRET = "medingen_secret_123"

ZOHO_CLIENT_ID = "1000.UR1N0QFKVQANWZSVSXOTAGW83KNLJN"
ZOHO_CLIENT_SECRET = "bab2f821ee9fe1d795c13fe294940f51f8967e9571"

def get_mysql_connection():
    return pymysql.connect(
        host=app.config['MYSQL_HOST'],
        user=app.config['MYSQL_USER'],
        password=app.config['MYSQL_PASSWORD'],
        db=app.config['MYSQL_DB'],
        cursorclass=pymysql.cursors.DictCursor  # Return results as dictionaries
    )

@app.route('/api/health', methods=['GET', 'POST'])
def health():
        return jsonify({'success': "success", "region": "ap-south-1"}), 200



def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None

        # Check if 'Authorization' header is present
        if 'Authorization' in request.headers:
            token = request.headers['Authorization'].split()[1]

        if not token:
            return jsonify({'error': 'Token is missing'}), 401

        try:
            # Decode the JWT token
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])

            if "user" in data.keys():
                current_user = data['user']
                connection = get_mysql_connection()
                with connection.cursor() as cursor:
                    cursor.execute("SELECT * FROM AdminUsers WHERE user_id = %s", (current_user,))
                    admin_user = cursor.fetchone()
                connection.close()

                if not admin_user:
                    return jsonify({'message': 'Admin user not found!'}), 403  # Forbidden

                # Check if the request has a customer query parameter
                print(request.args)
                customer_id = request.args.get('customer')
                if customer_id:
                    # Return the customer_id instead of admin_user if present
                    print("in to", customer_id)
                    current_user = customer_id

            elif "customer_id" in data.keys():
                current_user = data['customer_id']
                
        except Exception as e:
            print(e)
            return jsonify({'error': 'Token is invalid'}), 401

        return f(current_user, *args, **kwargs)

    return decorated



def create_notification_to_admin(sender, notification_message, action_url, noti_type="info", reminder_date=None):
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:

            safe_noti_type = noti_type if noti_type else "info"
            safe_reminder = reminder_date if reminder_date else None

            print("ADMIN NOTIFICATION DEBUG:",
                  sender, notification_message, action_url, safe_noti_type)

            for admin in ["1011"]:
                insert_query = """
                    INSERT INTO Notifications (
                        notification_receiver, notification_message, 
                        notification_sender, action_url, noti_type, reminder_date
                    )
                    VALUES (%s, %s, %s, %s, %s, %s)
                """
                cursor.execute(
                    insert_query,
                    (admin, notification_message, sender, action_url, safe_noti_type, safe_reminder)
                )

            mysql.commit()

            # ---------------------------------------------------------
            # for admin in ["1011", "1012"]:
            #     cursor.execute(
            #         "SELECT endpoint, p256dh, auth FROM PushSubscriptions WHERE user_id = %s",
            #         (admin,)
            #     )
            #     subscriptions = cursor.fetchall()
            #
            #     for subscription in subscriptions:
            #         api_data = {
            #             "payload": {
            #                 "title": "Medingen",
            #                 "body": notification_message,
            #                 "icon": "/android-chrome-192x192.png",
            #                 "target_url": action_url
            #             },
            #             "subscription": subscription,
            #             "schedule_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            #         }
            #
            #         api_url = "http://ec2-35-154-224-159.ap-south-1.compute.amazonaws.com/send_notification_247832438"
            #         response = requests.post(api_url, json=api_data)
            #         print(f"Notification sent, status code: {response.status_code}")
            # ---------------------------------------------------------

    except Exception as e:
        print("NOTIFICATION ERROR:", str(e))
        print(traceback.format_exc())

    finally:
        mysql.close()



@app.route('/api/create_order', methods=['POST'])
def create_order():
    data = request.get_json()
    amount = data["total_amount"]

    api_url = f"http://ec2-35-154-224-159.ap-south-1.compute.amazonaws.com/rzp_453432sxw?total_amount={amount*100}"
    response = requests.get(api_url)

    if response.status_code != 200:
        return jsonify({'message': 'Failed to create payment order'}), 500
    else:
        response_dict = response.json()
        rzp_order_id = response_dict.get("id")

        order_response = Response(
            response.content,
            status=response.status_code,
            content_type=response.headers.get('Content-Type')
        )        

    mysql = get_mysql_connection()
    try:
        data = request.get_json()
        coupon_savings = data.get("coupon_savings", 0)
        cart_id = data.get("cart_id", "")

        with mysql.cursor() as cursor:
            update_query = """
                UPDATE Cart
                SET rzp_order_id = %s, coupon_savings = %s
                WHERE cart_id = %s 
            """
            cursor.execute(update_query, (rzp_order_id, coupon_savings, cart_id))
            mysql.commit()

        return order_response
    except Exception as e:
        print(e)
        return jsonify({'message': 'Failed to update cart'}), 500
    finally:
        mysql.close()


@app.route('/api/verify_payment', methods=['POST'])
def verify_payment():
    try:
        payload = request.data
        razorpay_signature = request.headers.get('X-Razorpay-Signature')

        print("Headers:", dict(request.headers))
        print("Raw Payload:", payload)

        try:
            payload_json = json.loads(payload)
        except:
            payload_json = {}

        event_type = payload_json.get("event")
        print("Event Type:", event_type)

        if not razorpay_signature:
            return jsonify({"error": "Missing Razorpay signature"}), 400

        generated_signature = hmac.new(
            RAZORPAY_SECRET.encode(),
            payload,
            hashlib.sha256
        ).hexdigest()

        if not hmac.compare_digest(generated_signature, razorpay_signature):
            print("Invalid signature")
            return jsonify({"status": "failure", "message": "Invalid signature"}), 400

        if event_type == "payment.captured":
            print("Handling: PAYMENT CAPTURED (Order Flow)")

            payment_entity = payload_json["payload"]["payment"]["entity"]

            payment_id = payment_entity.get("id")
            order_id = payment_entity.get("order_id")
            amount = (payment_entity.get("amount") or 0) / 100

            print("Payment ID:", payment_id)
            print("Order ID:", order_id)

            if not order_id:
                print("Missing order_id; cannot process order payment.")
                return jsonify({"error": "missing order_id"}), 400

            mysql = get_mysql_connection()
            try:
                # Update Cart with payment status
                with mysql.cursor() as cursor:
                    cursor.execute("""
                        UPDATE Cart
                        SET cart_status = 'payment',
                            payment_id = %s,
                            payment_sign = %s,
                            payment_done_date = NOW()
                        WHERE rzp_order_id = %s;
                    """, (payment_id, razorpay_signature, order_id))
                    mysql.commit()

                # Log webhook event
                with mysql.cursor() as cursor:
                    cursor.execute("""
                        UPDATE Cart
                        SET payment_webhook_logs = JSON_ARRAY_APPEND(
                            IFNULL(payment_webhook_logs, JSON_ARRAY()),
                            '$',
                            JSON_OBJECT(
                                'status', 'SUCCESS',
                                'type', 'order_payment',
                                'payment_id', %s,
                                'order_id', %s,
                                'amount', %s,
                                'timestamp', NOW()
                            )
                        )
                        WHERE rzp_order_id = %s;
                    """, (payment_id, order_id, amount, order_id))
                    mysql.commit()

                with mysql.cursor(pymysql.cursors.DictCursor) as cursor:
                    cursor.execute("""
                        SELECT cart_id, customer_id
                        FROM Cart
                        WHERE rzp_order_id = %s
                        LIMIT 1;
                    """, (order_id,))
                    cart = cursor.fetchone()

                if cart:
                    cart_id = cart["cart_id"]
                    customer_id = cart["customer_id"]

                    try:
                        message = f"Payment received for order {cart_id}"
                        action_link = f"/admin/customer_detail/{customer_id}"

                        create_notification_to_admin(
                            sender=customer_id,
                            notification_message=message,
                            action_url=action_link,
                            noti_type="payment"
                        )

                        print("Admin payment notification saved.")

                    except Exception as notify_err:
                        print("Admin payment notify error:", notify_err)

            finally:
                mysql.close()

            return jsonify({"status": "success", "message": "Order payment verified"}), 200

        print("Unhandled event:", event_type)
        return jsonify({"status": "ignored", "event": event_type}), 200

    except Exception as e:
        print("Unexpected error:", e)
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/payment_link_webhook', methods=['POST'])
def payment_link_webhook():
    try:
        payload = request.data
        razorpay_signature = request.headers.get("X-Razorpay-Signature")

        print("Headers:", dict(request.headers))
        print("Raw Payload:", payload)

        try:
            data = json.loads(payload)
        except:
            data = {}

        event_type = data.get("event")
        print("Payment Link Webhook Event:", event_type)

        if not razorpay_signature:
            return jsonify({"error": "Missing signature"}), 400

        generated_signature = hmac.new(
            RAZORPAY_SECRET.encode(),
            payload,
            hashlib.sha256
        ).hexdigest()

        if not hmac.compare_digest(generated_signature, razorpay_signature):
            print("Invalid signature for payment link webhook")
            return jsonify({"error": "Invalid signature"}), 400

        link_entity = (
            data.get("payload", {})
                .get("payment_link", {})
                .get("entity", {})
        )

        plink_id = link_entity.get("id")
        plink_status = link_entity.get("status")
        plink_amount = (link_entity.get("amount") or 0) / 100
        plink_expire_by = link_entity.get("expire_by")
        reference_id = link_entity.get("reference_id")

        if not reference_id:
            return jsonify({"error": "reference_id missing"}), 400

        cart_id = reference_id

        print("Payment Link:", plink_id, "| Cart:", cart_id, "| Status:", plink_status)

        mysql = get_mysql_connection()

        # --------------------------------------------------------
        # UPDATE CART TABLE
        # --------------------------------------------------------
        with mysql.cursor() as cursor:
            cursor.execute("""
                UPDATE Cart
                SET 
                    rzp_plink_id = %s,
                    rzp_plink_status = %s,
                    payment_status = %s,
                    payment_expiry = FROM_UNIXTIME(%s),
                    rzp_plink_raw = %s,
                    cart_status = CASE WHEN %s = 'paid' THEN 'payment' ELSE cart_status END,
                    payment_done_date = CASE WHEN %s = 'paid' THEN NOW() ELSE payment_done_date END
                WHERE cart_id = %s
            """, (
                plink_id,
                plink_status,
                plink_status,
                plink_expire_by,
                json.dumps(data),
                plink_status,
                plink_status,
                cart_id
            ))

            # ----------------------------------------------------
            # UPDATE DISPATCHED TABLE (purchased_on)
            # ----------------------------------------------------
            if plink_status == "paid":
                cursor.execute("""
                    UPDATE Dispatched
                    SET purchased_on = NOW()
                    WHERE cart_id = %s
                """, (cart_id,))

            mysql.commit()

        # --------------------------------------------------------
        # LOG WEBHOOK EVENT
        # --------------------------------------------------------
        with mysql.cursor() as cursor:
            cursor.execute("""
                UPDATE Cart
                SET payment_webhook_logs = JSON_ARRAY_APPEND(
                    IFNULL(payment_webhook_logs, JSON_ARRAY()),
                    '$',
                    JSON_OBJECT(
                        'plink_id', %s,
                        'status', %s,
                        'amount', %s,
                        'event', %s,
                        'timestamp', NOW()
                    )
                )
                WHERE cart_id = %s
            """, (
                plink_id,
                plink_status,
                plink_amount,
                event_type,
                cart_id
            ))

            mysql.commit()

        # --------------------------------------------------------
        # FETCH customer_id FOR ADMIN NOTIFICATION
        # --------------------------------------------------------
        customer_id = None
        with mysql.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("""
                SELECT customer_id
                FROM Cart
                WHERE cart_id = %s
                LIMIT 1
            """, (cart_id,))
            row = cursor.fetchone()

            if row:
                customer_id = row["customer_id"]

        # --------------------------------------------------------
        # SEND ADMIN NOTIFICATION IF PAID
        # --------------------------------------------------------
        if plink_status == "paid" and customer_id:
            try:
                message = f"Payment link paid for cart {cart_id}"
                action_url = f"/admin/customer_detail/{customer_id}"

                create_notification_to_admin(
                    sender=customer_id,
                    notification_message=message,
                    action_url=action_url,
                    noti_type="payment"
                )

                print("Admin notification sent with customer_id:", customer_id)

            except Exception as e:
                print("Admin notification error:", e)

        mysql.close()
        return jsonify({"status": "success"}), 200

    except Exception as e:
        print("Webhook Error:", str(e))
        return jsonify({"error": str(e)}), 500

@app.route('/api/check_payment', methods=['POST'])
def check_payment():
    print("API Call: /api/check_payment")

    data = request.get_json(force=True)
    print(f"Incoming data: {data}")

    cart_id = data.get('cart_id')
    rzp_order_id = data.get('razorpay_order_id')
    rzp_payment_id = data.get('razorpay_payment_id')

    print(f"Extracted - cart_id: {cart_id}, rzp_order_id: {rzp_order_id}, rzp_payment_id: {rzp_payment_id}")

    if not all([cart_id, rzp_order_id]):
        print("Missing required fields")
        return jsonify({"status": "failure", "message": "Missing required fields"}), 400

    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            cursor.execute("""
              SELECT cart_id, rzp_order_id, payment_id, cart_status
              FROM Cart
              WHERE cart_id=%s AND rzp_order_id=%s
              LIMIT 1
            """, (cart_id, rzp_order_id))
            row = cursor.fetchone()

        if not row:
            return jsonify({"status": "not_found", "message": "Order not found for this cart"}), 404

        # Fill missing payment_id (fallback)
        if not row['payment_id'] and rzp_payment_id:
            with mysql.cursor() as cursor:
                cursor.execute("""
                    UPDATE Cart
                    SET payment_id=%s
                    WHERE cart_id=%s AND rzp_order_id=%s
                """, (rzp_payment_id, cart_id, rzp_order_id))

                cursor.execute("""
                    UPDATE Dispatched
                    SET purchased_on=NOW()
                    WHERE cart_id=%s
                """, (cart_id,))

            mysql.commit()
            row['payment_id'] = rzp_payment_id

        elif not row['payment_id']:
            return jsonify({"status": "pending", "message": "Payment not captured/verified yet"}), 202

        if rzp_payment_id and rzp_payment_id != row['payment_id']:
            return jsonify({"status": "mismatch", "message": "Different payment_id on record"}), 409

        return jsonify({
            "status": "paid",
            "message": "Payment recorded",
            "cart_status": row.get('cart_status'),
            "payment_id": row.get('payment_id')
        }), 200

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

    finally:
        mysql.close()


def generate_random_digits(n):
    """Generate a random number with n digits."""
    return ''.join(["{}".format(random.randint(0, 9)) for num in range(0, n)])



@app.route('/api/generate_presigned_url', methods=['GET'])
@token_required
def generate_presigned_url(current_user):
    try:
        file_name = request.args.get('file_name')
        prefix = request.args.get("prefix")
        random_str = request.args.get("random", "true").lower()
        random_gen = random_str in ["true", "1", "yes"]

        if not file_name:
            return jsonify({'error': 'File name is required'}), 400

        if random_gen:
            random_number = generate_random_digits(5)
            file_name = random_number + "_" + file_name

        # Generate a presigned URL for the S3 bucket
        presigned_url = s3_client.generate_presigned_url(
            'put_object',
            Params={
                'Bucket': BUCKET_NAME,
                'Key': prefix + "/" + file_name,
                'ContentType': request.args.get('content_type', 'application/octet-stream')            },
            ExpiresIn=3600  # URL expiration time in seconds
        )

        return jsonify({'presigned_url': presigned_url, "file_name": file_name}), 200
    except NoCredentialsError:
        return jsonify({'error': 'Credentials not available'}), 500
    except PartialCredentialsError:
        return jsonify({'error': 'Incomplete credentials provided'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route("/api/googleauthsignup", methods=["POST"])
def googleauthsignup():
    data = request.json
    phone_number = data.get('phone_number')
    token = data.get("token")
    otp = data.get('otp')

    # Connect to the database
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            # Verify OTP and retrieve user details
            query = """
                SELECT customer_id, customer_name
                FROM Customers 
                WHERE phonenumber = %s AND otp = %s
            """
            cursor.execute(query, (phone_number, otp))
            user = cursor.fetchone()

            if not user:
                return jsonify({'error': 'Phone number or OTP invalid'}), 404
            
            api_url = f"http://ec2-35-154-224-159.ap-south-1.compute.amazonaws.com/googleauth_452343"
            response = requests.post(api_url, json={"token": token})

            if response.status_code == 200:
                resp = response.json()
                
                user_id = resp["sub"]
                email = resp["email"]
                name = resp["name"]
                picture = resp["picture"]

            # Update the password
            update_query = "UPDATE Customers SET google_auth = %s, email = %s, customer_name = %s, profile_picture = %s WHERE customer_id = %s"
            cursor.execute(update_query, (user_id, email ,name, picture, user['customer_id']))
            mysql.commit()

            return jsonify({'message': 'Login updated successfully', "customer_name": name}), 200

    except Exception as e:
        print(e)
        return jsonify({'error': str(e)}), 500

    finally:
        # Ensure the database connection is properly closed
        mysql.close()

@app.route("/api/googleauth", methods=["POST"])
def google_auth():
    token = request.get_json().get("token")
    
    try:
        api_url = f"http://ec2-35-154-224-159.ap-south-1.compute.amazonaws.com/googleauth_452343"
        response = requests.post(api_url, json={"token": token})

        # Check if the request was successful
        if response.status_code == 200:
            resp = response.json()
            
            user_id = resp["sub"]
            email = resp["email"]
            name = resp["name"]
            picture = resp["picture"]

                # Query database to verify credentials
            mysql = get_mysql_connection()
            with mysql.cursor() as cursor:
                query = "SELECT * FROM Customers WHERE email = %s AND google_auth = %s"
                cursor.execute(query, (email,user_id))
                user = cursor.fetchone()

                if user:
                    # Update last login date
                    update_query = "UPDATE Customers SET last_login_date = %s WHERE customer_id = %s"
                    cursor.execute(update_query, (datetime.now(), user['customer_id']))
                    mysql.commit()

                    # Generate JWT token
                    token = jwt.encode({'customer_id': user['customer_id']}, app.config['SECRET_KEY'], algorithm="HS256")
                    
                    return jsonify({
                        'token': token,
                        'customer_id': user['customer_id'],
                        'customer_name': user['customer_name'],
                        'email': user['email']
                    }), 200
                else:
                    return jsonify({'error': 'User not signed up via google'}), 401
        else:
            return jsonify({"error": "Failed to authenticate user"}), 400

    except Exception as e:
        return jsonify({"error": str(e)}), 400



@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    hashed_password = data.get('hashedPassword')  

    # Query database to verify credentials
    mysql = get_mysql_connection()
    with mysql.cursor() as cursor:
        query = "SELECT * FROM AdminUsers WHERE username = %s AND password_hash = %s"
        cursor.execute(query, (username, hashed_password))
        user = cursor.fetchone()

        if user:
            update_query = "UPDATE AdminUsers SET last_login_date = %s WHERE user_id = %s"
            cursor.execute(update_query, (datetime.now(), user['user_id']))
            mysql.commit()
            token = jwt.encode({'user': user["user_id"]}, app.config['SECRET_KEY'], algorithm="HS256")
            return jsonify({'token': token, 'username': user["username"], 'email': user["email"]}), 200
        else:
            return jsonify({'error': 'Invalid credentials'}), 401



@app.route('/api/cart_count', methods=['GET'])
@token_required
def getCartCount(customer_id):
    from collections import Counter

    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            select_query = """
                SELECT product_ids_qty 
                FROM Cart 
                WHERE customer_id = %s 
                AND cart_status IN ('active','pending_confirm','confirm')
            """
            cursor.execute(select_query, (customer_id,))
            result = cursor.fetchone()

            product_ids_qty = result.get("product_ids_qty") if result else ""

            total_count = 0
            combined_cart_items = ""

            if product_ids_qty:
                raw_items = [itm for itm in product_ids_qty.split(';') if itm]
                product_counter = Counter()

                for item in raw_items:
                    if ':' not in item:
                        continue

                    product_id, quantity = item.split(':', 1)

                    try:
                        qty = int(quantity)
                    except:
                        qty = 0

                    product_counter[product_id] += qty
                    total_count += qty

                combined_cart_items = ";".join(
                    f"{pid}:{qty}" for pid, qty in product_counter.items()
                )

        return jsonify({
            "message": "Cart details",
            "cart_items": combined_cart_items,
            "cart_count": total_count
        }), 200

    except Exception as e:
        raise e
    finally:
        mysql.close()


@app.route('/api/update_delivery_address', methods=['POST'])
@token_required
def update_delivery_address(customer_id):
    mysql = get_mysql_connection()
    try:
        data = request.get_json()
        address_id = data.get('address_id')
        cart_id = data.get("cart_id")

        if not address_id:
            return jsonify({'message': 'Address ID is required'}), 400
        
        if not cart_id:
             return jsonify({'message': 'Cart ID is required'}), 400

        with mysql.cursor() as cursor:
            # Get the pincode for the selected address
            cursor.execute("SELECT pincode FROM Addresses WHERE id = %s", (address_id,))
            addr = cursor.fetchone()
            
            computed_shipping = 50
            if addr and addr.get('pincode'):
                pin = str(addr.get('pincode')).strip()
                if pin and pin[0] in ['5', '6']:
                    computed_shipping = 50
                elif pin and pin[0] in ['1', '2', '3', '4', '7', '8']:
                    computed_shipping = 60
                else:
                    computed_shipping = 50 # Current default

            # Update the delivery address ID and shipping charge
            update_query = """
                UPDATE Cart
                SET delivery_address_id = %s,
                    shipping_charge = %s
                WHERE customer_id = %s
                AND cart_id = %s
                AND cart_status IN ('active', 'pending_confirm', 'confirm')
            """
            cursor.execute(update_query, (
                address_id,
                computed_shipping,
                customer_id,
                cart_id
            ))
            mysql.commit()

        return jsonify({'message': 'Delivery address updated successfully'}), 200
    except Exception as e:
        print(e)
        return jsonify({'message': 'Failed to update delivery address'}), 500
    finally:
        mysql.close()



@app.route('/api/update_choose_prescription', methods=['POST'])
@token_required
def update_choose_prescription(customer_id):
    mysql = get_mysql_connection()
    try:
        data = request.get_json()
        prescription_id = data.get('prescription_id')
        cart_id = data.get("cart_id", "cart_id")

        if not prescription_id:
            return jsonify({'message': 'Prescription ID is required'}), 400

        with mysql.cursor() as cursor:
            # Update the prescription id
            update_query = """
                UPDATE Cart
                SET prescription_id = %s
                WHERE customer_id = %s
                AND cart_status IN ('active', 'pending_confirm', 'confirm')
                AND cart_id = %s
            """
            cursor.execute(update_query, (
                prescription_id,
                customer_id,
                cart_id
            ))
            mysql.commit()

        return jsonify({'message': 'Prescription updated successfully'}), 200
    except Exception as e:
        print("Failed to update delivery Prescription",e)
        return jsonify({'message': 'Failed to update delivery Prescription'}), 500
    finally:
        mysql.close()


@app.route('/api/remove_prescription_from_cart', methods=['POST'])
@token_required
def remove_prescription_from_cart(customer_id):
    mysql = get_mysql_connection()
    try:
        data = request.get_json()
        cart_id = data.get("cart_id")

        if not cart_id:
            return jsonify({'message': 'Cart ID is required'}), 400

        with mysql.cursor() as cursor:
            # Update the prescription id to 0 (remove)
            update_query = """
                UPDATE Cart
                SET prescription_id = 0
                WHERE customer_id = %s
                AND cart_id = %s
                AND cart_status IN ('active', 'pending_confirm', 'confirm')
            """
            cursor.execute(update_query, (customer_id, cart_id))
            mysql.commit()

        return jsonify({'message': 'Prescription removed successfully'}), 200
    except Exception as e:
        print("Failed to remove Prescription", e)
        return jsonify({'message': 'Failed to remove Prescription'}), 500
    finally:
        mysql.close()


@app.route('/api/update_delivery_charge', methods=['POST'])
@token_required
def update_delivery_charge(customer_id):
    mysql = get_mysql_connection()
    try:
        data = request.get_json()
        cart_id = data.get("cart_id", "cart_id")
        shipping_charge = data.get("shipping_charge", 50)


        with mysql.cursor() as cursor:
            # Update cart with new quantities
            update_query = """
                UPDATE Cart
                SET shipping_charge=%s
                WHERE cart_id = %s
            """
            cursor.execute(update_query, (shipping_charge, cart_id ))
            mysql.commit()

        return jsonify({'message': 'Cart updated successfully'}), 200
    except Exception as e:
        print(e)
        return jsonify({'message': 'Failed to update cart'}), 500
    finally:
        mysql.close()


@app.route("/api/cart/payment-update", methods=["POST"])
def update_cart_payment():
    mysql = get_mysql_connection()

    try:
        data = request.get_json(silent=True) or {}

        cart_id = data.get("cart_id")
        cart_status = data.get("cart_status")
        payment_mode = data.get("payment_mode")

        if not cart_id:
            return jsonify({"message": "cart_id is required"}), 400

        # ✅ added online
        allowed_modes = ["cod", "scanner", "online"]

        if payment_mode and payment_mode.lower() not in allowed_modes:
            return jsonify({
                "message": "Only 'cod', 'scanner', and 'online' allowed"
            }), 400

        with mysql.cursor() as cursor:

            cursor.execute(
                "SELECT payment_mode, customer_id FROM Cart WHERE cart_id = %s",
                (cart_id,)
            )
            cart = cursor.fetchone()

            if not cart:
                return jsonify({"message": "Cart not found"}), 404

            final_mode = (payment_mode or cart["payment_mode"] or "").lower()
            customer_id = cart["customer_id"]

            fields = []
            values = []

            # payment_mode update
            if payment_mode:
                fields.append("payment_mode = %s")
                values.append(final_mode)

            # status update for all valid modes
            if cart_status and final_mode in allowed_modes:
                fields.append("cart_status = %s")
                values.append(cart_status.lower())

            payment_date = None

            # ✅ COD + ONLINE → mark payment done
            if final_mode in ["cod", "online"]:
                payment_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                fields.append("payment_done_date = %s")
                values.append(payment_date)

            if not fields:
                return jsonify({"message": "No valid fields"}), 400

            values.append(cart_id)

            query = f"""
                UPDATE Cart
                SET {", ".join(fields)}
                WHERE cart_id = %s
            """

            cursor.execute(query, tuple(values))

            # ---------------- DISPATCHED ----------------
            if payment_date:
                cursor.execute(
                    "SELECT dispatched_id FROM Dispatched WHERE cart_id = %s",
                    (cart_id,)
                )
                dispatched = cursor.fetchone()

                if dispatched:
                    cursor.execute("""
                        UPDATE Dispatched
                        SET purchased_on = %s
                        WHERE cart_id = %s
                    """, (payment_date, cart_id))
                else:
                    cursor.execute("""
                        INSERT INTO Dispatched (cart_id, purchased_on)
                        VALUES (%s, %s)
                    """, (cart_id, payment_date))

            mysql.commit()

            # ---------------- NOTIFICATION ----------------
            if final_mode in ["cod", "online"] and payment_date:

                cursor.execute("""
                    SELECT 1 FROM Notifications
                    WHERE notification_message LIKE %s
                    LIMIT 1
                """, (f"%Cart ID {cart_id}%",))

                exists = cursor.fetchone()

                if not exists:
                    create_notification_to_admin(
                        sender="system",
                        notification_message=f"Payment received for order {cart_id}",
                        action_url=f"/customer_detail/{customer_id}",
                        noti_type="payment"
                    )

            return jsonify({
                "message": "Cart updated successfully",
                "payment_done_date": payment_date
            }), 200

    except Exception as e:
        mysql.rollback()
        return jsonify({"message": str(e)}), 500

    finally:
        mysql.close()

FREEZE_STATUSES = ['confirm', 'dispatched', 'delivered', 'payment', 'cancelled', 'return', 'refund', 'paid', 'success']

# ---------------- SNAPSHOT ----------------
def create_price_snapshot(product_ids_qty, mysql):

    items = []

    if not product_ids_qty:
        return items

    parsed_items = []

    for item in product_ids_qty.split(';'):

        item = item.strip()

        if not item or ':' not in item:
            continue

        try:
            product_id, qty = item.split(':')

            parsed_items.append({
                "product_id": int(product_id),
                "qty": int(qty)
            })

        except:
            continue

    if not parsed_items:
        return items

    product_ids = [x["product_id"] for x in parsed_items]

    format_strings = ','.join(['%s'] * len(product_ids))

    query = f"""
        SELECT
            product_id,
            name,
            product_pricing_new,
            product_pricing_old,
            rc
        FROM Products
        WHERE product_id IN ({format_strings})
    """

    with mysql.cursor() as cursor:

        print("FETCHING PRODUCTS")

        cursor.execute(query, product_ids)

        products = cursor.fetchall()

    product_map = {
        p['product_id']: p
        for p in products
    }

    for row in parsed_items:

        product = product_map.get(row['product_id'])

        if not product:
            continue

        price = float(product['product_pricing_new'] or 0)
        mrp = float(product['product_pricing_old'] or 0)
        rc_val = float(product['rc'] or 0)

        items.append({
            "product_id": row['product_id'],
            "name": product['name'],
            "qty": row['qty'],
            "price": price,
            "mrp": mrp,
            "rc": rc_val,
            "total": price * row['qty']
        })

    return items


       
@app.route('/api/cart_update', methods=['POST'])
@token_required
def update_cart(customer_id):
    mysql = get_mysql_connection()
    try:
        data = request.get_json()
        product_quantities = data.get('quantities', {})
        cart_id = data.get("cart_id", "")

        if product_quantities is None:
            return jsonify({'message': 'No quantities provided'}), 400

        if not cart_id:
            return jsonify({'message': 'Cart ID is required'}), 400

        validated_quantities = {}

        # ✅ Handle dictionary format
        if isinstance(product_quantities, dict):
            for product_id, qty in product_quantities.items():
                try:
                    qty = int(qty)

                    if qty < 0:
                        return jsonify({
                            'message': f'Quantity for product {product_id} cannot be negative'
                        }), 400

                    # ✅ keep 0 also
                    validated_quantities[str(product_id)] = qty

                except (ValueError, TypeError):
                    return jsonify({
                        'message': f'Invalid quantity for product {product_id}'
                    }), 400

        # ✅ Handle string format
        elif isinstance(product_quantities, str):
            for item in product_quantities.split(';'):
                if not item or ':' not in item:
                    continue
                try:
                    product_id, qty = item.split(':')
                    qty = int(qty)

                    if qty < 0:
                        return jsonify({
                            'message': f'Quantity for product {product_id} cannot be negative'
                        }), 400

                    # ✅ keep 0 also
                    validated_quantities[str(product_id)] = qty

                except (ValueError, TypeError):
                    return jsonify({
                        'message': 'Invalid format in product quantities'
                    }), 400
        else:
            return jsonify({'message': 'Invalid quantities format'}), 400

        # ✅ Always allow empty or zero quantities
        if validated_quantities:
            product_ids_qty_string = ";".join(
                f"{pid}:{qty}" for pid, qty in validated_quantities.items()
            ) + ";"
        else:
            product_ids_qty_string = ""

        action_url = f"/admin/customer_detail/{customer_id}"
        message = f"Cart updated for cart {cart_id}"
        customerId = f"{customer_id}"

        with mysql.cursor() as cursor:

            # ✅ Verify cart
            verify_query = """
                SELECT cart_id, cart_status FROM Cart
                WHERE cart_id = %s AND customer_id = %s
            """
            cursor.execute(verify_query, (cart_id, customer_id))
            cart_data = cursor.fetchone()

            if not cart_data:
                return jsonify({
                    'message': 'Cart not found or does not belong to you'
                }), 404

            if cart_data['cart_status'] not in ('active', 'pending_confirm', 'confirm'):
                return jsonify({
                    'message': f'Cart is already {cart_data["cart_status"]} and cannot be updated'
                }), 404

            # ✅ Update cart
            update_query = """
                UPDATE Cart
                SET product_ids_qty = %s,
                    cart_status='active',
                    cart_updated_date=NOW()
                WHERE cart_id = %s AND customer_id = %s
            """
            cursor.execute(update_query, (product_ids_qty_string, cart_id, customer_id))
            mysql.commit()

            # ✅ Notifications
            check_query = """
                SELECT notification_id FROM Notifications
                WHERE action_url = %s AND noti_type = 'order'
                LIMIT 1
            """
            cursor.execute(check_query, (action_url,))
            existing = cursor.fetchone()

            if existing:
                update_notif = """
                    UPDATE Notifications
                    SET notification_message = %s,
                        date_received = NOW(),
                        read_status = NULL,
                        is_read = 0
                    WHERE notification_id = %s
                """
                cursor.execute(update_notif, (message, existing["notification_id"]))
            else:
                insert_notif = """
                    INSERT INTO Notifications
                    (notification_sender, notification_receiver, notification_message,
                     date_received, action_url, noti_type, is_read, created_at)
                    VALUES (%s, 1011, %s, NOW(), %s, 'order', 0, NOW())
                """
                cursor.execute(insert_notif, (customerId, message, action_url))

            mysql.commit()

        return jsonify({
            'message': 'Cart updated successfully',
            'updated_quantities': validated_quantities
        }), 200

    except Exception as e:
        print(f"Error: {e}")
        mysql.rollback()
        return jsonify({'message': 'Failed to update cart'}), 500

    finally:
        mysql.close()


@app.route('/api/place_order', methods=['POST'])
@token_required
def place_order(customer_id):
    mysql = get_mysql_connection()
    try:
        data = request.get_json()
        cart_id = data.get("cart_id", "cart_id")
        total_cart_value = data.get("total_cart_value", 0)
        delivery_type = data.get("delivery_type", "normal")

        action_url = f"/admin/customer_detail/{customer_id}"
        message = f"Cart updated for cart {cart_id}"
        customerId = f"{customer_id}"

        with mysql.cursor() as cursor:

            update_query = """
                UPDATE Cart 
                SET cart_status = "pending_confirm", 
                    pending_confirm_at = NOW(),
                    total_cart_value = %s,
                    delivery_type = %s,
                    delivery_address_id = IF(delivery_address_id = 0, 
                                            (SELECT id FROM Addresses WHERE customer_id = %s AND `default` = 1 LIMIT 1), 
                                            delivery_address_id)
                WHERE cart_id = %s
            """
            cursor.execute(update_query, (total_cart_value, delivery_type, customer_id, cart_id))
            mysql.commit()

            check_query = """
                SELECT notification_id FROM Notifications
                WHERE action_url = %s AND noti_type = 'order'
                LIMIT 1
            """
            cursor.execute(check_query, (action_url,))
            existing = cursor.fetchone()

            if existing:
                update_notif = """
                    UPDATE Notifications
                    SET notification_message = %s, 
                        date_received = NOW(),
                        read_status = NULL,
                        is_read = 0
                    WHERE notification_id = %s
                """
                cursor.execute(update_notif, (message, existing["notification_id"]))
            else:
                insert_notif = """
                    INSERT INTO Notifications 
                    (notification_sender, notification_receiver, notification_message, 
                    date_received, action_url, noti_type, is_read, created_at)
                    VALUES (%s, 1011, %s, NOW(), %s, 'order', 0, NOW())
                """
                cursor.execute(insert_notif, (customerId, message, action_url))

            mysql.commit()

        return jsonify({'message': 'Cart updated successfully'}), 200

    except Exception as e:
        print(e)
        return jsonify({'message': 'Failed to update cart'}), 500

    finally:
        mysql.close()


@app.route('/api/re_active_cart', methods=['POST'])
@token_required
def re_active_cart(customer_id):
    mysql = get_mysql_connection()
    try:
        data = request.get_json()
        cart_id = data.get("cart_id")
        if not cart_id:
            return jsonify({'message': 'Cart ID is required'}), 400
            
        with mysql.cursor() as cursor:
            update_query = "UPDATE Cart SET cart_status = 'active' WHERE cart_id = %s"
            cursor.execute(update_query, (cart_id,))
            mysql.commit()

        return jsonify({'message': 'Cart status updated to active'}), 200
    except Exception as e:
        print(e)
        return jsonify({'message': 'Failed to reactivate cart'}), 500
    finally:
        mysql.close()


@app.route('/api/update_cod_charge', methods=['POST'])
@token_required
def update_cod_charge(current_user):
    mysql = get_mysql_connection()
    try:
        data = request.get_json()
        cart_id = data.get("cart_id")
        cod_charge = data.get("cod_charge", 0)

        if not cart_id:
            return jsonify({'message': 'cart_id is required'}), 400

        with mysql.cursor() as cursor:
            update_query = """
                UPDATE Cart
                SET cod_charge=%s
                WHERE cart_id = %s
            """
            cursor.execute(update_query, (cod_charge, cart_id))
            mysql.commit()

        return jsonify({'message': 'COD charge updated successfully'}), 200
    except Exception as e:
        print(e)
        return jsonify({'message': 'Failed to update COD charge'}), 500
    finally:
        mysql.close()



@app.route('/api/add-to-cart', methods=['POST'])
@token_required
def add_to_cart(customer_id):
    data = request.get_json()
    product_id = data.get('product_id')
    quantity = data.get('quantity')
    prescription_id = data.get('prescription_id')

    # Validate quantity
    try:
        qty = int(quantity) if quantity else 0
        if qty < 0:
            return jsonify({"error": "Quantity cannot be negative"}), 400
        if qty == 0 and not product_id:
            return jsonify({"error": "Quantity must be greater than 0"}), 400
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid quantity"}), 400

    # normalize product
    if product_id:
        pid = product_id
    else:
        pid = 38743
        qty = 0

    mysql = get_mysql_connection()

    try:
        with mysql.cursor() as cursor:

            cursor.execute(
                "SELECT cart_id FROM Cart WHERE customer_id=%s AND cart_status IN ('active','pending_confirm','confirm')",
                (customer_id,)
            )
            cart = cursor.fetchone()

            if cart:
                cart_id = cart['cart_id']

                if prescription_id:
                    cursor.execute(
                        """UPDATE Cart 
                           SET product_ids_qty = CONCAT(IFNULL(product_ids_qty,''), %s, ':', %s, ';'),
                               cart_status='active',
                               prescription_id=%s
                           WHERE cart_id=%s""",
                        (pid, qty, prescription_id, cart_id)
                    )
                else:
                    cursor.execute(
                        """UPDATE Cart 
                           SET product_ids_qty = CONCAT(IFNULL(product_ids_qty,''), %s, ':', %s, ';'),
                               cart_status='active'
                           WHERE cart_id=%s""",
                        (pid, qty, cart_id)
                    )

            else:
                if prescription_id:
                    cursor.execute(
                        """INSERT INTO Cart (customer_id, cart_status, product_ids_qty, prescription_id)
                           VALUES (%s,'active',%s,%s)""",
                        (customer_id, f"{pid}:{qty};", prescription_id)
                    )
                else:
                    cursor.execute(
                        """INSERT INTO Cart (customer_id, cart_status, product_ids_qty)
                           VALUES (%s,'active',%s)""",
                        (customer_id, f"{pid}:{qty};")
                    )
                cart_id = cursor.lastrowid

            cursor.execute("SELECT product_ids_qty FROM Cart WHERE cart_id=%s", (cart_id,))
            updated_count = cursor.fetchone()['product_ids_qty']

            action_url = f"/admin/customer_detail/{customer_id}"
            message = f"Cart updated for cart {cart_id}"
            customerId = f"{customer_id}"

            cursor.execute(
                """SELECT notification_id FROM Notifications
                   WHERE action_url=%s AND noti_type='order'
                   LIMIT 1""",
                (action_url,)
            )
            existing = cursor.fetchone()

            if existing:
                cursor.execute(
                    """UPDATE Notifications
                       SET notification_message=%s,
                           date_received=NOW(),
                           read_status=NULL,
                           is_read=0
                       WHERE notification_id=%s""",
                    (message, existing["notification_id"])
                )
            else:
                cursor.execute(
                    """INSERT INTO Notifications
                       (notification_sender, notification_receiver, notification_message,
                        date_received, action_url, noti_type, is_read, created_at)
                       VALUES (%s, 1011, %s, NOW(), %s, 'order', 0, NOW())""",
                    (customerId, message, action_url)
                )

            mysql.commit()

        return jsonify({"message": "Product added to cart successfully", "cart_items": updated_count}), 200

    except Exception as e:
        mysql.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()


def get_product_ids_from_cart(cart):
    product_ids_qty = cart['product_ids_qty']
    product_ids = [item.split(":")[0] for item in product_ids_qty.split(";") if item]
    return product_ids

@app.route('/api/cart_coupons/<int:cart_id>', methods=['GET'])
def get_applicable_coupons(cart_id):
    conn = get_mysql_connection()
    try:
        with conn.cursor() as cursor:
            # Fetch the cart
            cursor.execute("SELECT * FROM Cart WHERE cart_id = %s", (cart_id,))
            cart = cursor.fetchone()

            if not cart:
                return jsonify({'message': 'Cart not found'}), 404
            
            # Extract product IDs from the cart
            product_ids_in_cart = get_product_ids_from_cart(cart)

            # Fetch all active coupons
            cursor.execute("SELECT * FROM Coupons WHERE coupon_status = 'active'")
            all_coupons = cursor.fetchall()

            available_coupons = []
            unavailable_coupons = []
            
            # Filter applicable and unavailable coupons
            for coupon in all_coupons:
                applicable_products = [p.strip() for p in coupon['applicable_on_product_ids'].split(",") if p.strip()]
                # If coupon is applicable to all products or contains products in the cart
                if "*" in applicable_products or any(product_id in applicable_products for product_id in product_ids_in_cart):
                    # Check if the total cart value meets the minimum order value for the coupon
                    if cart['total_cart_value'] >= float(coupon['minimum_order_value']):
                        available_coupons.append(coupon)
                    else:
                        unavailable_coupons.append(coupon)
                else:
                    unavailable_coupons.append(coupon)

            return jsonify({
                'available_coupons': available_coupons,
                'unavailable_coupons': unavailable_coupons
            })
    finally:
        conn.close()

@app.route('/api/apply_coupon', methods=['POST'])
@token_required
def apply_coupon(customer_id):
    conn = get_mysql_connection()
    data = request.get_json()
    cart_id = data.get('cart_id')
    coupon_code = data.get('coupon_code')

    try:
        with conn.cursor() as cursor:
            # Fetch the cart
            cursor.execute("SELECT * FROM Cart WHERE cart_id = %s", (cart_id,))
            cart = cursor.fetchone()

            if not cart:
                return jsonify({'message': 'Cart not found'}), 404
            
            # Extract product IDs from the cart
            product_ids_in_cart = get_product_ids_from_cart(cart)

            # Fetch all active coupons
            cursor.execute("SELECT * FROM Coupons WHERE coupon_status = 'active' AND coupon_code=%s", (coupon_code,) )
            all_coupons = cursor.fetchall()

            available_coupons = []
            unavailable_coupons = []
            
            # Filter applicable and unavailable coupons
            for coupon in all_coupons:
                applicable_products = [p.strip() for p in coupon['applicable_on_product_ids'].split(",") if p.strip()]
                # If coupon is applicable to all products or contains products in the cart
                if "*" in applicable_products or any(product_id in applicable_products for product_id in product_ids_in_cart):
                    # Check if the total cart value meets the minimum order value for the coupon
                    if cart['total_cart_value'] >= float(coupon['minimum_order_value']):
                        available_coupons.append(coupon)
                    else:
                        unavailable_coupons.append(coupon)
                else:
                    unavailable_coupons.append(coupon)

            if len(available_coupons):
                coupon = available_coupons[0]
                if coupon["coupon_type"] == "percentage":
                    coupon_savings = float(cart['total_cart_value']) * (float(coupon['discount_value']) /100)
                elif coupon["coupon_type"] == "fixed" or coupon["coupon_type"] == None :
                    coupon_savings = coupon['discount_value']
                else:
                    coupon_savings = 0
                return jsonify({'coupon_savings': round(coupon_savings, 2)}), 200
            else:
                return jsonify({'statusText': 'Coupon code not applicable'}), 404
    except Exception as e:
        print(e)
        return jsonify({'statusText': 'Backend error'}), 500
    finally:
        conn.close()


DTDC_BUCKET = "medingen-store-new"
DTDC_KEY = "pincode/dtdc.xlsx"
DTDC_REGION = "ap-south-1"

_DTDc_PINS = None
_LAST_LOAD = 0
_CACHE_TTL = 600


def _load_dtdc_pincodes():
    global _DTDc_PINS, _LAST_LOAD

    now = time.time()
    if _DTDc_PINS is not None and (now - _LAST_LOAD) < _CACHE_TTL:
        return _DTDc_PINS

    try:
        s3 = boto3.client("s3", region_name=DTDC_REGION)

        obj = s3.get_object(Bucket=DTDC_BUCKET, Key=DTDC_KEY)
        data = obj["Body"].read()

        wb = load_workbook(BytesIO(data), read_only=True)
        ws = wb.worksheets[0]

        header = next(ws.iter_rows(max_row=1, values_only=True))
        pin_idx = next(
            (i for i, h in enumerate(header) if h and "PINCODE" in str(h).upper()),
            None
        )

        if pin_idx is None:
            _DTDc_PINS = set()
            return _DTDc_PINS

        pins = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            pin = row[pin_idx]
            if pin is not None:
                pins.add(str(int(pin)))

        wb.close()

        _DTDc_PINS = pins
        _LAST_LOAD = now
        return _DTDc_PINS

    except Exception:
        _DTDc_PINS = set()
        return _DTDc_PINS


def is_dtdc_available(pincode):
    if not pincode:
        return False
    return str(pincode).strip() in _load_dtdc_pincodes()

@app.route('/api/cart', methods=['GET'])
@token_required
def get_cart(customer_id):
    start_time = time.time()

    mysql = get_mysql_connection()
    all_orders = int(request.args.get('all_orders', 0))
    cart_id = int(request.args.get('cart_id', -1))
    all_carts = []

    try:
        with mysql.cursor(pymysql.cursors.DictCursor) as cursor:

            cursor.execute("""
                SELECT phonenumber, customer_name, total_mig_coins_available
                FROM Customers
                WHERE customer_id = %s
            """, (customer_id,))
            customer_info = cursor.fetchone() or {}

            customer_phone = customer_info.get("phonenumber")
            customer_name = customer_info.get("customer_name")
            total_mig_coins = customer_info.get("total_mig_coins_available", 0)

            cart_query = """
                SELECT cart_status, payment_mode, product_ids_qty,
                       delivery_address_id, prescription_id, cart_id,
                       shipping_charge, cod_charge, cart_updated_date,
                       cart_created_date, order_tracking_id,
                       payment_id, coupon_savings, coupon_applied,
                       delivery_type, offer_id, payment_done_date,
                       is_followed, whatsapp_response,
                       pending_confirm_at
                FROM Cart
                WHERE customer_id = %s
            """

            if not all_orders and cart_id <= 0:
                cart_query += " AND cart_status IN ('active','confirm','pending_confirm')"

            if cart_id > 0:
                cart_query += " AND cart_id = %s"

            cart_query += " ORDER BY cart_updated_date DESC"

            params = (customer_id, cart_id) if cart_id > 0 else (customer_id,)
            cursor.execute(cart_query, params)
            carts = cursor.fetchall()

            # ✅ FIX 1: return 200 when no carts
            if not carts:
                print("Execution time:", time.time() - start_time)
                return jsonify([] if all_orders else {}), 200

            all_product_ids = set()
            cart_products_map = {}

            for cart in carts:
                product_quantities = {}
                for pq in (cart.get("product_ids_qty") or "").split(";"):
                    if ":" in pq:
                        pid, qty = pq.split(":")
                        try:
                            qty = int(qty)
                            product_quantities[pid] = product_quantities.get(pid, 0) + qty
                            all_product_ids.add(pid)
                        except:
                            pass
                cart_products_map[cart["cart_id"]] = product_quantities

            products_dict = {}
            if all_product_ids:
                placeholders = ",".join(["%s"] * len(all_product_ids))
                cursor.execute(
                    f"""
                    SELECT product_id, name, product_name_url,
                           manufacturer, product_pricing_old,
                           product_pricing_new,
                           IFNULL(rc,0) AS rc,
                           IFNULL(inventory_info_total_stock,0) AS product_available,
                           JSON_UNQUOTE(JSON_EXTRACT(photo,'$[0].img')) AS first_image_url,
                           prescription_required
                    FROM Products
                    WHERE product_id IN ({placeholders})
                    """,
                    tuple(all_product_ids)
                )
                for p in cursor.fetchall():
                    products_dict[str(p["product_id"])] = p

            cursor.execute("""
                SELECT id, name, address1 AS addressLine1,
                       state AS addressLine2, state,
                       pincode, type, phone_number,
                       whatsapp_response, `default`
                FROM Addresses
                WHERE customer_id = %s
            """, (customer_id,))
            all_addresses = cursor.fetchall()
            address_map = {a["id"]: a for a in all_addresses}
            default_address = next((a for a in all_addresses if a.get("default") == 1), None)

            cursor.execute("SELECT id, title FROM Offers")
            offers = {o["id"]: o["title"] for o in cursor.fetchall()}

            prescription_ids = list({c["prescription_id"] for c in carts if c["prescription_id"]})
            prescriptions = {}
            if prescription_ids:
                placeholders = ",".join(["%s"] * len(prescription_ids))
                cursor.execute(
                    f"""
                    SELECT prescription_id, prescription_image_url,
                           prescription_date, prescription_status,
                           prescription_comments, last_used_date,
                           prescription_name
                    FROM Prescription
                    WHERE prescription_id IN ({placeholders})
                    """,
                    tuple(prescription_ids)
                )
                prescriptions = {p["prescription_id"]: p for p in cursor.fetchall()}

            for cart in carts:

                product_quantities = cart_products_map.get(cart["cart_id"], {})
                products_data = []
                total_mrp = 0
                total_selling_price = 0

                for pid, qty in product_quantities.items():
                    p = products_dict.get(pid)
                    if not p:
                        continue

                    old_price = p["product_pricing_old"] or 0
                    new_price = p["product_pricing_new"] or 0

                    total_mrp += old_price * qty
                    total_selling_price += new_price * qty

                    products_data.append({
                        "id": p["product_id"],
                        "image": p["first_image_url"],
                        "name": p["name"],
                        "product_name_url": p["product_name_url"],
                        "manufacturer": p["manufacturer"],
                        "originalPrice": f"Rs. {old_price}",
                        "discountedPrice": f"Rs. {new_price}",
                        "rc": int(p["rc"]),
                        "inStock": int(p["product_available"]) > 0,
                        "quantity": qty,
                        "discountPercentage": (
                            f"{round(((old_price - new_price) / old_price) * 100)}% less Price"
                            if old_price > 0 and new_price > 0 else "0%"
                        ),
                        "prescriptionRequired": p["prescription_required"]
                    })

                shipping_charge = Decimal(cart.get("shipping_charge") or 0)
                cod_charge = Decimal(cart.get("cod_charge") or 0)

                expected_offer_id = 0
                if total_selling_price >= 1000:
                    expected_offer_id = 2
                elif total_selling_price >= 500:
                    expected_offer_id = 1

                if cart.get("offer_id") != expected_offer_id:
                    with mysql.cursor() as write_cursor:
                        write_cursor.execute(
                            "UPDATE Cart SET offer_id = %s WHERE cart_id = %s",
                            (expected_offer_id, cart["cart_id"])
                        )
                        mysql.commit()
                    cart["offer_id"] = expected_offer_id

                total_payable = Decimal(str(total_selling_price)) + shipping_charge + cod_charge

                if cart.get("offer_id") == 2 and total_payable > 1000:
                    total_payable -= (total_payable * Decimal("0.05")).quantize(Decimal("0.01"))

                all_carts.append({
                    "cart": products_data,
                    "cart_id": cart["cart_id"],
                    "is_followed": bool(cart.get("is_followed")),
                    "delivery_type": cart["delivery_type"],
                    "customer_phone": customer_phone,
                    "customer_name": customer_name,
                    "orderSummary": {
                        "itemsCount": sum(product_quantities.values()),
                        "totalMRP": f"Rs. {total_mrp}",
                        "total_selling_price": f"Rs. {total_selling_price}",
                        "totalPercentageSaved": (
                            f"{round(((total_mrp - total_selling_price) / total_mrp) * 100)}%"
                            if total_mrp > 0 else "0%"
                        ),
                        "total_shipping_charge": float(shipping_charge),
                        "cod_charge": float(cod_charge),
                        "totalSavings": f"Rs. {total_mrp - total_selling_price}",
                        "migCoins": total_mig_coins,
                        "totalAmount": float(total_payable),
                    },
                    "cart_updated_date": cart["cart_updated_date"],
                    "cart_created_date": cart["cart_created_date"],
                    "payment_done_date": cart["payment_done_date"],
                    "pending_confirm_at": cart.get("pending_confirm_at"),
                    "deliveryAddress": address_map.get(cart["delivery_address_id"]) or default_address,
                    "cartStatus": cart["cart_status"],
                    "paymentmode": cart["payment_mode"],
                    "order_tracking_id": cart["order_tracking_id"],
                    "payment_id": cart["payment_id"],
                    "coupon_savings": cart["coupon_savings"],
                    "coupon_applied": cart["coupon_applied"],
                    "offerTitle": offers.get(cart["offer_id"]),
                    "prescriptionDetails": prescriptions.get(cart["prescription_id"], {}),
                    "whatsapp_response": cart.get("whatsapp_response"),
                    "courier_tracking": []
                })

            print("Execution time:", time.time() - start_time)

            # ✅ FIX 2: safe return (avoid IndexError)
            if not all_carts:
                return jsonify([] if all_orders else {}), 200

            return jsonify(all_carts if all_orders else all_carts[0]), 200

    finally:
        mysql.close()



@app.route("/api/cart/<int:cart_id>/assign-offer", methods=["POST"])
def assign_offer(cart_id):
    data = request.get_json()
    offer_id = data.get("offer_id")

    try:
        conn = get_mysql_connection()
        cursor = conn.cursor()

        if offer_id in (None, "", "none"):
            cursor.execute("UPDATE Cart SET offer_id = NULL WHERE cart_id = %s", (cart_id,))
            conn.commit()
            return jsonify({
                "message": "Offer removed successfully",
                "cart_id": cart_id,
                "offer": None
            }), 200

        # Validate if offer exists
        cursor.execute("SELECT * FROM Offers WHERE id = %s", (offer_id,))
        offer = cursor.fetchone()
        if not offer:
            return jsonify({"error": "Invalid offer_id"}), 404

        # Update cart with selected offer
        cursor.execute(
            "UPDATE Cart SET offer_id = %s WHERE cart_id = %s",
            (offer_id, cart_id)
        )
        conn.commit()

        return jsonify({
            "message": "Offer assigned successfully",
            "cart_id": cart_id,
            "offer": offer
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()
        conn.close()


@app.route('/api/customers-by-type', methods=['GET'])
@token_required
def get_customers_by_type(current_user):
    from datetime import datetime, date
    import traceback

    customer_type = request.args.get("customer_type")
    if not customer_type:
        return jsonify({"success": False, "error": "customer_type is required"}), 400

    try:
        mysql = get_mysql_connection()

        with mysql.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("""
                SELECT
                    customer_id,
                    customer_name,
                    email,
                    phonenumber,
                    customer_category,
                    customer_type
                FROM Customers
                WHERE customer_type = %s
                ORDER BY customer_id DESC
            """, (customer_type,))
            customers = cursor.fetchall()

        if not customers:
            return jsonify({"success": True, "customers": [], "count": 0}), 200

        customer_ids = [c["customer_id"] for c in customers]

        with mysql.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("""
                SELECT
                    c.cart_id,
                    c.customer_id,
                    c.cart_status,
                    c.payment_mode,
                    c.product_ids_qty,
                    d.dispatched_id,
                    d.dispatched_date,
                    d.delivered_date,
                    d.reminder_on,
                    ROW_NUMBER() OVER (
                        PARTITION BY c.customer_id
                        ORDER BY c.cart_updated_date DESC
                    ) rn
                FROM Cart c
                LEFT JOIN Dispatched d ON d.cart_id = c.cart_id
                WHERE c.customer_id IN %s
            """, (tuple(customer_ids),))
            carts = cursor.fetchall()

        carts = [c for c in carts if c["rn"] <= 5]

        product_ids = set()
        for c in carts:
            raw = c.get("product_ids_qty") or ""
            for x in raw.split(";"):
                if ":" in x:
                    pid, qty = x.split(":")
                    if qty.isdigit() and int(qty) > 0:
                        product_ids.add(pid)

        product_lookup = {}
        if product_ids:
            placeholders = ",".join(["%s"] * len(product_ids))
            with mysql.cursor(pymysql.cursors.DictCursor) as cursor:
                cursor.execute(f"""
                    SELECT product_id, name, product_name_url
                    FROM Products
                    WHERE product_id IN ({placeholders})
                """, tuple(product_ids))
                for p in cursor.fetchall():
                    product_lookup[str(p["product_id"])] = p

        carts_by_customer = {}

        for c in carts:
            cid = c["customer_id"]
            carts_by_customer.setdefault(cid, [])

            products = []
            raw = c.get("product_ids_qty") or ""
            for x in raw.split(";"):
                if ":" in x:
                    pid, qty = x.split(":")
                    try:
                        qty = int(qty)
                    except:
                        qty = 0
                    if qty > 0:
                        prod = product_lookup.get(pid)
                        products.append({
                            "product_id": pid,
                            "name": prod["name"] if prod else f"Product {pid}",
                            "product_name_url": prod.get("product_name_url") if prod else None,
                            "qty": qty
                        })

            for f in ["dispatched_date", "delivered_date", "reminder_on"]:
                if isinstance(c.get(f), (datetime, date)):
                    c[f] = c[f].isoformat()

            carts_by_customer[cid].append({
                "cart_id": c["cart_id"],
                "dispatched_id": c.get("dispatched_id"),
                "cart_status": c["cart_status"],
                "payment_mode": c["payment_mode"],
                "products": products,
                "dispatched_date": c.get("dispatched_date"),
                "delivered_date": c.get("delivered_date"),
                "reminder_on": c.get("reminder_on")
            })

        for cust in customers:
            cust["carts"] = carts_by_customer.get(cust["customer_id"], [])

        return jsonify({
            "success": True,
            "customers": customers,
            "count": len(customers)
        }), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        mysql.close()






@app.route('/api/request-product', methods=['POST'])
@token_required
def request_product(current_user):
    data = request.get_json()
    if not data:
        return jsonify({"message": "No data provided"}), 400

    product_id = data.get("product_id")
    customer_id = data.get("customer_id")
    prescription_id = data.get("prescription_id")
    status = data.get("status", "pending").lower()
    mode = data.get("mode", "Insert").lower()

    if not all([product_id, customer_id]):
        return jsonify({"message": "Missing required fields"}), 400

    try:
        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            if mode == "update" and prescription_id:
                update_query = """
                UPDATE ProductRequest
                SET status = %s
                WHERE prescription_id = %s
                """
                cursor.execute(update_query, (status, prescription_id))
                mysql.commit()
                return jsonify({"message": "Product request updated successfully"}), 200
            else:
                insert_query = """
                INSERT INTO ProductRequest (product_id, customer_id, request_date, status, prescription_id)
                VALUES (%s, %s, NOW(), %s, %s)
                """
                cursor.execute(insert_query, (
                    product_id,
                    customer_id,
                    status,
                    prescription_id 
                ))
                mysql.commit()
                request_id = cursor.lastrowid

                create_notification_to_admin(
                    customer_id,
                    f"Product request for product {product_id}",
                    f"/admin/product-edit/{product_id}",
                    "product_request"
                )

                return jsonify({
                    "message": "Product request created successfully",
                    "request_id": request_id
                }), 200

    except Exception as e:
        print("Error handling ProductRequest:", e)
        return jsonify({"message": "An error occurred while processing product request"}), 500


@app.route('/api/products', methods=['POST'])
def get_products():
    data = request.json
    page = int(data.get('page', 1))
    query_string = data.get('query', '')
    text = data.get('text', '')
    show_hidden = data.get('show_hidden', False)

    category = None
    if ":" in text:
        category, text = text.split(":", 1)
        if text == "":
            text = " "

    per_page = 10
    offset = (page - 1) * per_page

    base_query = """
        FROM Products p
        LEFT JOIN (
            SELECT DISTINCT category_name, category_outline_url
            FROM Category
        ) c ON p.categories = c.category_name
        WHERE 1=1
    """

    params = []

    if query_string:
        base_query += f" AND ({query_string})"

    if text:
        normalized_text = text.replace('-', '').replace(' ', '')
        like_pattern = f"%{normalized_text}%"

        base_query += """
            AND (
                REPLACE(REPLACE(p.name, '-', ''), ' ', '') LIKE %s
                OR REPLACE(REPLACE(p.salt_name, '-', ''), ' ', '') LIKE %s
                OR REPLACE(REPLACE(p.composition, '-', ''), ' ', '') LIKE %s
            )
        """
        params.extend([like_pattern, like_pattern, like_pattern])

        if category:
            base_query += " AND p.categories = %s"
            params.append(category)

    if not show_hidden:
        base_query += " AND p.visibility_status = 'Published'"

    mysql = get_mysql_connection()

    # ✅ Total count
    count_query = "SELECT COUNT(DISTINCT p.product_id) AS total_count " + base_query

    with mysql.cursor() as cursor:
        cursor.execute(count_query, params)
        total_count = cursor.fetchone()['total_count']

    # ✅ Main select
    select_query = """
        SELECT 
            p.name AS product_name, 
            p.product_id, 
            JSON_UNQUOTE(JSON_EXTRACT(p.photo, '$[0].img')) AS first_image_url,
            p.product_pricing_new, 
            p.product_pricing_old,
            p.salt_name, 
            p.composition,
            p.categories as selected_category, 
            p.inventory_info_total_stock as product_available,
            True as product_request,
            p.rc,
            LENGTH(p.composition) as lencomp,
            p.product_name_url,
            c.category_outline_url
    """ + base_query

    if "rc=0" in query_string:
        pricing_order = "DESC"
    else:
        pricing_order = "ASC"

    if text:
        select_query += f"""
            GROUP BY p.product_id
            ORDER BY 
                p.rc DESC,
                LOCATE(%s, p.name) DESC,
                lencomp,
                p.product_pricing_new {pricing_order},
                p.product_id
            LIMIT %s OFFSET %s
        """
        final_params = params + [text, per_page, offset]
    else:
        select_query += f"""
            GROUP BY p.product_id
            ORDER BY 
                p.product_pricing_new {pricing_order}
            LIMIT %s OFFSET %s
        """
        final_params = params + [per_page, offset]

    with mysql.cursor() as cursor:
        cursor.execute(select_query, final_params)
        products = cursor.fetchall()

    return jsonify({
        "results": products,
        "total_count": total_count
    }), 200

@app.route('/api/salt_products', methods=['POST'])
def get_products_salt():
    data = request.json or {}

    page = int(data.get('page', 1))
    text = data.get('text', '')
    show_hidden = data.get('show_hidden', False)
    rc = data.get('rc', None)

    per_page = 10
    offset = (page - 1) * per_page

    category = None
    if ":" in text:
        category, text = text.split(":", 1)
        if text == "":
            text = " "

    query = """
        SELECT 
            p.name AS product_name, 
            p.product_id, 
            JSON_UNQUOTE(JSON_EXTRACT(p.photo, '$[0].img')) AS first_image_url,
            p.product_pricing_new, 
            p.product_pricing_old,
            p.salt_name, 
            p.composition,
            p.categories AS selected_category, 
            p.inventory_info_total_stock AS product_available,
            TRUE AS product_request,
            p.rc,
            LENGTH(p.composition) AS lencomp,
            p.product_name_url,
            MAX(c.category_outline_url) AS category_outline_url
        FROM Products p 
        LEFT JOIN Category c 
            ON p.categories = c.category_name
        WHERE 1=1
    """

    params = []

    if text:
        normalized_text = text.replace('-', '').replace(' ', '')
        query += """
            AND (
                REPLACE(REPLACE(p.name, '-', ''), ' ', '') LIKE %s
                OR REPLACE(REPLACE(p.salt_name, '-', ''), ' ', '') LIKE %s
                OR REPLACE(REPLACE(p.composition, '-', ''), ' ', '') LIKE %s
            )
        """
        like_pattern = f"%{normalized_text}%"
        params.extend([like_pattern, like_pattern, like_pattern])

        if category:
            query += " AND p.categories = %s "
            params.append(category)

    if rc is not None:
        query += " AND p.rc = %s "
        params.append(rc)

    if not show_hidden:
        query += " AND p.visibility_status = 'Published' "

    # 👇 KEY FIX — REMOVE DUPLICATES
    query += " GROUP BY p.product_id "

    mysql = get_mysql_connection()

    # total count (after grouping)
    with mysql.cursor() as cursor:
        cursor.execute(query, params)
        total_count = len(cursor.fetchall())

    pricing_order = "ASC"
    if rc == 0:
        pricing_order = "DESC"

    if text:
        query += f"""
            ORDER BY p.rc DESC, lencomp ASC, p.product_pricing_new {pricing_order}, p.product_id
        """
    else:
        query += f"""
            ORDER BY p.product_pricing_new {pricing_order}
        """

    query += " LIMIT %s OFFSET %s "
    params.extend([per_page, offset])

    with mysql.cursor() as cursor:
        cursor.execute(query, params)
        products = cursor.fetchall()

    return jsonify({"results": products, "total_count": total_count}), 200


@app.route('/api/alt_products', methods=['POST'])
def get_products_alt():
    try:
        data = request.json or {}

        page = int(data.get('page', 1))
        per_page = 10
        offset = (page - 1) * per_page

        composition = data.get('composition', '')
        exclude_product_id = data.get('exclude_product_id', None)
        rc = int(data.get('rc', 1))
        show_hidden = data.get('show_hidden', False)

        mysql = get_mysql_connection()

        base_query = """
            FROM Products p
            LEFT JOIN (
                SELECT DISTINCT category_name, category_outline_url
                FROM Category
            ) c ON p.categories = c.category_name
            WHERE 1=1
        """

        params = []

        if composition:
            base_query += " AND LOWER(TRIM(p.composition)) = LOWER(TRIM(%s)) "
            params.append(composition)

        if exclude_product_id:
            base_query += " AND p.product_id != %s "
            params.append(exclude_product_id)

        base_query += " AND p.rc = %s "
        params.append(rc)

        if not show_hidden:
            base_query += " AND p.visibility_status = 'Published' "

        base_query += " AND p.product_pricing_new IS NOT NULL "

        count_query = "SELECT COUNT(DISTINCT p.product_id) AS total " + base_query

        with mysql.cursor() as cursor:
            cursor.execute(count_query, params)
            total_count = cursor.fetchone()['total']

        final_query = f"""
            SELECT
                p.name AS product_name,
                p.product_id,
                p.photo AS images,
                p.product_pricing_new,
                p.product_pricing_old,
                p.salt_name,
                p.composition,
                p.categories AS selected_category,
                p.inStock,
                p.rc,
                p.packaging,
                p.manufacturer,
                p.product_name_url,
                c.category_outline_url
            {base_query}
            GROUP BY p.product_id
            ORDER BY p.product_pricing_new ASC
            LIMIT %s OFFSET %s
        """

        final_params = params + [per_page, offset]

        with mysql.cursor() as cursor:
            cursor.execute(final_query, final_params)
            results = cursor.fetchall()

        total_pages = (total_count + per_page - 1) // per_page

        return jsonify({
            'current_page': page,
            'total_pages': total_pages,
            'total_results': total_count,
            'results': results
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/get_products_by_category', methods=['POST'])
def get_products_by_category():
    data = request.get_json()

    category_id = data.get('category_id')
    category_name = data.get('category_name')

    if category_name:
        category_name = category_name.replace('_', ' ')

    page = int(data.get('page', 1))
    per_page = int(data.get('per_page', 12))
    sort_by = data.get('sort_by', 'price_low_high')

    consume_type = data.get('consume_type')
    composition = data.get('composition')
    sub_categories = data.get('sub_categories', [])

    if not category_id and not category_name:
        return jsonify({"message": "Category ID or Name required"}), 400

    offset = (page - 1) * per_page
    mysql = get_mysql_connection()

    try:
        with mysql.cursor(pymysql.cursors.DictCursor) as cursor:

            target_categories = []
            is_main_category = False
            main_cat_name = None

            conditions = []
            params = []

            if category_id:
                conditions.append("category_id = %s")
                params.append(category_id)

            if category_name:
                conditions.append("category_name = %s")
                params.append(category_name)

            cursor.execute(
                f"SELECT * FROM Category WHERE {' OR '.join(conditions)} LIMIT 1",
                tuple(params)
            )
            subcategory = cursor.fetchone()

            if subcategory:
                target_categories.append(subcategory['category_name'])
            else:
                conditions_main = []
                params_main = []

                if category_id:
                    conditions_main.append("id = %s")
                    params_main.append(category_id)

                if category_name:
                    conditions_main.append("name = %s")
                    params_main.append(category_name)

                cursor.execute(
                    f"SELECT * FROM MainCategory WHERE {' OR '.join(conditions_main)} LIMIT 1",
                    tuple(params_main)
                )
                main_category = cursor.fetchone()

                if not main_category:
                    return jsonify({"message": "Category not found"}), 404

                is_main_category = True
                main_cat_name = main_category['name']

                cursor.execute(
                    "SELECT category_name FROM Category WHERE main_category = %s",
                    (main_cat_name,)
                )
                subs = cursor.fetchall()

                if not subs:
                    return jsonify({"results": [], "total_count": 0}), 200

                if sub_categories and len(sub_categories) > 0:
                    target_categories = [
                        s for s in sub_categories
                        if s in [sub['category_name'] for sub in subs]
                    ]
                    if not target_categories:
                        target_categories = sub_categories
                else:
                    target_categories = [s['category_name'] for s in subs]

            placeholders = ",".join(["%s"] * len(target_categories))

            # ---------------- MAIN QUERY ----------------
            query = f"""
                SELECT 
                    p.name AS product_name,
                    p.product_id,
                    JSON_UNQUOTE(JSON_EXTRACT(p.photo, '$[0].img')) AS first_image_url,
                    p.product_pricing_new,
                    p.product_pricing_old,
                    p.salt_name,
                    p.composition,
                    p.consume_type,
                    p.categories,
                    p.inventory_info_total_stock,
                    p.rc,
                    p.product_name_url,
                    c.category_outline_url,
                    ((p.product_pricing_old - p.product_pricing_new) 
                      / p.product_pricing_old) * 100 AS discount_percent
                FROM Products p
                LEFT JOIN Category c ON p.categories = c.category_name
                WHERE p.visibility_status = 'Published'
                AND p.categories IN ({placeholders})
                AND p.product_name_url IS NOT NULL
                AND p.product_name_url != ''
            """

            query_params = list(target_categories)

            if consume_type:
                query += " AND p.consume_type = %s"
                query_params.append(consume_type)

            if composition:
                query += " AND p.composition LIKE %s"
                query_params.append(f"%{composition}%")

            if sort_by == "price_low_high":
                query += " ORDER BY p.product_pricing_new ASC "
            elif sort_by == "price_high_low":
                query += " ORDER BY p.product_pricing_new DESC "
            elif sort_by == "discount_high_low":
                query += " ORDER BY discount_percent DESC "
            elif sort_by == "new_arrivals":
                query += " ORDER BY p.product_entry_created_date DESC "
            elif sort_by == "discount_low_high":
                query += " ORDER BY discount_percent ASC "
            else:
                query += " ORDER BY p.product_pricing_new ASC "

            query += " LIMIT %s OFFSET %s"
            query_params.extend([per_page, offset])

            cursor.execute(query, tuple(query_params))
            products = cursor.fetchall()

            # ---------------- COUNT QUERY ----------------
            count_query = f"""
                SELECT COUNT(*) as count
                FROM Products p
                WHERE p.visibility_status = 'Published'
                AND p.categories IN ({placeholders})
                AND p.product_name_url IS NOT NULL
                AND p.product_name_url != ''
            """

            count_params = list(target_categories)

            if consume_type:
                count_query += " AND p.consume_type = %s"
                count_params.append(consume_type)

            if composition:
                count_query += " AND p.composition LIKE %s"
                count_params.append(f"%{composition}%")

            cursor.execute(count_query, tuple(count_params))
            total_count = cursor.fetchone()['count']

            total_pages = (total_count + per_page - 1) // per_page

            return jsonify({
                "results": products,
                "current_page": page,
                "total_pages": total_pages,
                "total_results": total_count,
                "category_type": "main" if is_main_category else "sub",
                "main_category_name": main_cat_name
            }), 200

    except Exception as e:
        return jsonify({"message": str(e)}), 500

    finally:
        mysql.close()
        

@app.route('/api/avg_price', methods=['POST'])
def avg_price():
    data = request.json
    composition = data.get('composition', '')
    salt_name = data.get('salt_name', '')

    if composition:
        query = """
            SELECT AVG(product_pricing_old)  AS avg_product_pricing_old
                    FROM Products 
                    WHERE composition = %s 
        """
        search_term = composition
    else:
        query = """
            SELECT AVG(product_pricing_old)  AS avg_product_pricing_old
                    FROM Products 
                    WHERE salt_name = %s 
        """
        search_term = salt_name

    mysql = get_mysql_connection()

    with mysql.cursor() as cursor:
        cursor.execute(query, (search_term,))
        result = cursor.fetchone()

    response = {
        "averagePrice": result["avg_product_pricing_old"]
    }

    return jsonify(response)



def get_tracking_details(order_id):
    api_url = "http://ec2-35-154-224-159.ap-south-1.compute.amazonaws.com/order_tracking_34434523"
    response = requests.post(api_url, json={"order_id": order_id})
    
    if response.status_code == 200:
        try:
            return response.json()
        except ValueError:
            print("Invalid JSON response")
            return {}
    else:
        print(f"Error: status code {response.status_code}")
        return {}



def cancel_courier(order_id):

    api_url = f"http://ec2-35-154-224-159.ap-south-1.compute.amazonaws.com/order_cancel_34234342"
    response = requests.post(api_url, json={"order_id": order_id})
    print(response)

    if response.status_code == 200:
        resp = response.json()
        if resp.get("status", False) == "OK":
            for r in  resp.get("successConsignments", []):
                return r.get("success", False)
        else:
            return False
    else:
        return False

@app.route('/api/cancel_order', methods=['POST'])
@token_required
def cancel_order(customer_id):
    data = request.get_json()
    cart_id = data.get('cart_id')

    if not cart_id:
        return jsonify({'message': 'cart_id is required'}), 400

    mysql = get_mysql_connection()
    try:
        with mysql.cursor(pymysql.cursors.DictCursor) as cursor:

            cursor.execute(
                "SELECT * FROM Cart WHERE cart_id = %s AND customer_id = %s",
                (cart_id, customer_id)
            )
            cart = cursor.fetchone()

            if not cart:
                return jsonify({'message': 'Cart not found'}), 404

            # Cancel courier only if tracking exists
            if cart.get("order_tracking_id"):
                cancel_courier(cart["order_tracking_id"])

            cursor.execute(
                "UPDATE Cart SET cart_status = 'cancelled' WHERE cart_id = %s",
                (cart_id,)
            )
            mysql.commit()

            return jsonify({'message': 'Order cancelled successfully'}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({'message': 'Failed to cancel order'}), 500

    finally:
        mysql.close()

def format_date(value):
    if not value:
        return None
    if isinstance(value, str):
        try:
            # MySQL DATETIME default format
            return datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            return value
    try:
        return value.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(value)

@app.route('/api/orders', methods=['GET'])
@token_required
def get_orders(current_user):
    start_time = time.time()

    mysql = get_mysql_connection()

    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 3))
        offset = (page - 1) * per_page

        FREEZE_STATUSES = [
            'confirm','dispatched','delivered',
            'payment','cancelled','return','refund'
        ]

        with mysql.cursor(pymysql.cursors.DictCursor) as cursor:

            cursor.execute("""
                SELECT COUNT(*) AS total_count
                FROM Cart
                WHERE customer_id=%s
                AND cart_status!='active'
            """,(current_user,))

            total_count = cursor.fetchone()['total_count']

            cursor.execute("""
                SELECT phonenumber, customer_name, total_mig_coins_available
                FROM Customers
                WHERE customer_id=%s
            """,(current_user,))

            customer_info = cursor.fetchone() or {}

            customer_phone = customer_info.get("phonenumber")
            customer_name = customer_info.get("customer_name")
            total_mig_coins = customer_info.get("total_mig_coins_available",0)

            cursor.execute("""
                SELECT
                    c.cart_status,
                    c.payment_mode,
                    c.product_ids_qty,
                    c.delivery_address_id,
                    c.prescription_id,
                    c.cart_id,
                    c.shipping_charge,
                    c.cod_charge,
                    c.cart_updated_date,
                    c.cart_created_date,
                    c.order_tracking_id,
                    c.payment_id,
                    c.coupon_savings,
                    c.coupon_applied,
                    c.delivery_type,
                    c.offer_id,
                    c.payment_done_date,
                    c.is_followed,
                    c.whatsapp_response,
                    c.pending_confirm_at,
                    c.price_snapshot,

                    COALESCE(
                        NULLIF(d.tracking_id,''),
                        c.order_tracking_id
                    ) AS dispatched_tracking_id,

                    d.courier_type,
                    d.dispatched_date,
                    d.estimated_delivery,
                    d.delivered_date,
                    d.returned_date

                FROM Cart c
                LEFT JOIN Dispatched d
                ON d.cart_id = c.cart_id

                WHERE c.customer_id=%s
                AND c.cart_status!='active'

                ORDER BY c.cart_updated_date DESC
                LIMIT %s OFFSET %s
            """,(current_user,per_page,offset))

            carts = cursor.fetchall()

            if not carts:
                return jsonify({
                    "orders":[],
                    "total_pages":0,
                    "current_page":page,
                    "has_next":False,
                    "has_prev":False
                }),200

            all_product_ids = set()
            cart_products_map = {}

            for cart in carts:

                product_quantities = {}

                for pq in (cart.get("product_ids_qty") or "").split(";"):

                    if ":" in pq:
                        pid,qty = pq.split(":")

                        try:
                            qty = int(qty)

                            product_quantities[pid] = (
                                product_quantities.get(pid,0) + qty
                            )

                            all_product_ids.add(pid)

                        except:
                            pass

                if cart.get("price_snapshot"):

                    try:
                        snapshot_items = json.loads(
                            cart["price_snapshot"]
                        )

                        for item in snapshot_items:

                            if item.get("product_id"):
                                all_product_ids.add(
                                    str(item["product_id"])
                                )

                    except:
                        pass

                cart_products_map[cart["cart_id"]] = product_quantities

            products_dict = {}

            if all_product_ids:

                placeholders = ",".join(
                    ["%s"] * len(all_product_ids)
                )

                cursor.execute(f"""
                    SELECT
                        product_id,
                        name,
                        product_name_url,
                        manufacturer,
                        product_pricing_old,
                        product_pricing_new,
                        IFNULL(rc,0) AS rc,
                        IFNULL(inventory_info_total_stock,0) AS product_available,
                        JSON_UNQUOTE(JSON_EXTRACT(photo,'$[0].img')) AS first_image_url,
                        prescription_required
                    FROM Products
                    WHERE product_id IN ({placeholders})
                """,tuple(all_product_ids))

                for p in cursor.fetchall():
                    products_dict[str(p["product_id"])] = p

            cursor.execute("""
                SELECT
                    id,
                    name,
                    address1 AS addressLine1,
                    state AS addressLine2,
                    state,
                    pincode,
                    type,
                    phone_number,
                    whatsapp_response,
                    `default`
                FROM Addresses
                WHERE customer_id=%s
            """,(current_user,))

            all_addresses = cursor.fetchall()

            address_map = {
                a["id"]:a for a in all_addresses
            }

            default_address = next(
                (
                    a for a in all_addresses
                    if a.get("default") == 1
                ),
                None
            )

            cursor.execute("""
                SELECT
                    id,title
                FROM Offers
            """)

            offers = {
                o["id"]:o["title"]
                for o in cursor.fetchall()
            }

            prescription_ids = list({
                c["prescription_id"]
                for c in carts
                if c["prescription_id"]
            })

            prescriptions = {}

            if prescription_ids:

                placeholders = ",".join(
                    ["%s"] * len(prescription_ids)
                )

                cursor.execute(f"""
                    SELECT
                        prescription_id,
                        prescription_image_url,
                        prescription_date,
                        prescription_status,
                        prescription_comments,
                        last_used_date,
                        prescription_name
                    FROM Prescription
                    WHERE prescription_id IN ({placeholders})
                """,tuple(prescription_ids))

                prescriptions = {
                    p["prescription_id"]:p
                    for p in cursor.fetchall()
                }

            result = []

            for cart in carts:

                product_quantities = cart_products_map.get(
                    cart["cart_id"],{}
                )

                products_data = []

                total_mrp = 0
                total_selling_price = 0

                snapshot_items = []

                if cart.get("price_snapshot"):

                    try:
                        snapshot_items = json.loads(
                            cart["price_snapshot"]
                        )

                    except:
                        snapshot_items = []

                if (
                    snapshot_items and
                    cart["cart_status"] in FREEZE_STATUSES
                ):

                    for item in snapshot_items:

                        pid = str(item.get("product_id"))

                        db = products_dict.get(pid,{})

                        qty = int(item.get("qty",0))

                        mrp = float(item.get("mrp",0))
                        price = float(item.get("price",0))

                        total_mrp += mrp * qty
                        total_selling_price += price * qty

                        name = item.get("name") or db.get("name")

                        if str(name).strip().lower() == "prescription order":
                            continue

                        products_data.append({
                            "id":pid,
                            "image":item.get("image") or db.get("first_image_url"),
                            "name":name,
                            "product_name_url":item.get("product_name_url") or db.get("product_name_url"),
                            "manufacturer":item.get("manufacturer") or db.get("manufacturer"),
                            "originalPrice":f"Rs. {mrp}",
                            "discountedPrice":f"Rs. {price}",
                            "rc":int(db.get("rc",0)),
                            "inStock":int(db.get("product_available",0)) > 0,
                            "quantity":qty,
                            "discountPercentage":(
                                f"{round(((mrp-price)/mrp)*100)}% less Price"
                                if mrp > 0 and price > 0 else "0%"
                            ),
                            "prescriptionRequired":item.get(
                                "prescription_required",
                                db.get("prescription_required",0)
                            )
                        })

                else:

                    for pid,qty in product_quantities.items():

                        p = products_dict.get(pid)

                        if not p:
                            continue

                        old_price = p["product_pricing_old"] or 0
                        new_price = p["product_pricing_new"] or 0

                        total_mrp += old_price * qty
                        total_selling_price += new_price * qty

                        products_data.append({
                            "id":p["product_id"],
                            "image":p["first_image_url"],
                            "name":p["name"],
                            "product_name_url":p["product_name_url"],
                            "manufacturer":p["manufacturer"],
                            "originalPrice":f"Rs. {old_price}",
                            "discountedPrice":f"Rs. {new_price}",
                            "rc":int(p["rc"]),
                            "inStock":int(p["product_available"]) > 0,
                            "quantity":qty,
                            "discountPercentage":(
                                f"{round(((old_price-new_price)/old_price)*100)}% less Price"
                                if old_price > 0 and new_price > 0 else "0%"
                            ),
                            "prescriptionRequired":p["prescription_required"]
                        })

                shipping_charge = Decimal(
                    cart.get("shipping_charge") or 0
                )

                cod_charge = Decimal(
                    cart.get("cod_charge") or 0
                )

                total_payable = (
                    Decimal(str(total_selling_price))
                    + shipping_charge
                    + cod_charge
                )

                if (
                    cart.get("offer_id") == 2 and
                    total_payable > 1000
                ):
                    total_payable -= (
                        total_payable * Decimal("0.05")
                    ).quantize(Decimal("0.01"))

                courier_tracking = []

                if cart.get("dispatched_tracking_id"):
                    courier_tracking.append({
                        "tracking_id":cart.get("dispatched_tracking_id"),
                        "courier_type":cart.get("courier_type"),
                        "dispatched_date":str(cart.get("dispatched_date")) if cart.get("dispatched_date") else None,
                        "estimated_delivery":str(cart.get("estimated_delivery")) if cart.get("estimated_delivery") else None,
                        "delivered_date":str(cart.get("delivered_date")) if cart.get("delivered_date") else None,
                        "returned_date":str(cart.get("returned_date")) if cart.get("returned_date") else None
                    })

                result.append({
                    "cart":products_data,
                    "products":products_data,
                    "cart_id":cart["cart_id"],
                    "is_followed":bool(cart.get("is_followed")),
                    "delivery_type":cart["delivery_type"],
                    "customer_phone":customer_phone,
                    "customer_name":customer_name,
                    "orderSummary":{
                        "itemsCount":sum(product_quantities.values()),
                        "totalMRP":f"Rs. {total_mrp}",
                        "total_selling_price":f"Rs. {total_selling_price}",
                        "totalPercentageSaved":(
                            f"{round(((total_mrp-total_selling_price)/total_mrp)*100)}%"
                            if total_mrp > 0 else "0%"
                        ),
                        "total_shipping_charge":float(shipping_charge),
                        "cod_charge":float(cod_charge),
                        "totalSavings":f"Rs. {total_mrp-total_selling_price}",
                        "migCoins":total_mig_coins,
                        "totalAmount":float(total_payable)
                    },
                    "cart_updated_date":cart["cart_updated_date"],
                    "cart_created_date":cart["cart_created_date"],
                    "payment_done_date":cart["payment_done_date"],
                    "pending_confirm_at":cart.get("pending_confirm_at"),
                    "deliveryAddress":address_map.get(
                        cart["delivery_address_id"]
                    ) or default_address,
                    "cartStatus":cart["cart_status"],
                    "paymentmode":cart["payment_mode"],

                    "order_tracking_id":cart.get(
                        "dispatched_tracking_id"
                    ) or cart.get("order_tracking_id"),

                    "payment_id":cart["payment_id"],
                    "coupon_savings":cart["coupon_savings"],
                    "coupon_applied":cart["coupon_applied"],
                    "offerTitle":offers.get(cart["offer_id"]),
                    "prescriptionDetails":prescriptions.get(
                        cart["prescription_id"],{}
                    ),
                    "whatsapp_response":cart.get("whatsapp_response"),

                    "courier_tracking":courier_tracking
                })

        total_pages = (
            total_count + per_page - 1
        ) // per_page

        print("Execution time:",time.time() - start_time)

        return jsonify({
            "orders":result,
            "total_pages":total_pages,
            "current_page":page,
            "has_next":page < total_pages,
            "has_prev":page > 1
        }),200

    except Exception as e:

        return jsonify({
            "message":str(e)
        }),500

    finally:
        mysql.close()

        
def convert_to_timestamp(date_str, time_str):
    try:
        if len(date_str) != 8 or len(time_str) < 3:
            return 0
        time_str = time_str.zfill(4)
        dt = datetime.strptime(f"{date_str} {time_str}", "%d%m%Y %H%M")
        return int(dt.timestamp() * 1000)
    except Exception as e:
        return 0


@app.route('/api/order-tracking/<tracking_id>', methods=['POST'])
def get_order_tracking(tracking_id):
    try:
        tracking_raw = get_tracking_details(tracking_id)

        if not tracking_raw or not tracking_raw.get('trackDetails'):
            return jsonify({
                "tracking_id": tracking_id,
                "tracking": [],
                "trackHeader": {},
                "message": "No tracking details found"
            }), 200

        track_header = tracking_raw.get("trackHeader", {})
        latest_events = {}

        for detail in tracking_raw.get('trackDetails', []):
            event = detail.get("strAction", detail.get("mileName", "")).strip()
            location = detail.get("strOrigin", detail.get("mileLocationName", "")).strip()
            date = detail.get("strActionDate", detail.get("mileStatusDateTime", "")).split()[0]
            time = detail.get("strActionTime", detail.get("mileStatusDateTime", "")).split()[1] if " " in detail.get("strActionTime", detail.get("mileStatusDateTime", "")) else ""

            try:
                timestamp = convert_to_timestamp(date, time)
            except Exception:
                timestamp = 0

            key = (location, event)
            if key not in latest_events or timestamp > latest_events[key]['timestamp']:
                latest_events[key] = {
                    "location": location,
                    "timestamp": timestamp,
                    "event_description": event,
                    "status_external": event,
                    "status_internal": re.sub(r'[^a-z0-9_]', '', event.lower().replace(" ", "_")) if event else "unknown"
                }

        tracking_sorted = sorted(latest_events.values(), key=lambda x: x['timestamp'], reverse=True)

        booked_date = track_header.get("strBookedDate", "")
        booked_time = track_header.get("strBookedTime", "")
        booked_origin = track_header.get("strOrigin", "")

        try:
            booked_timestamp = convert_to_timestamp(booked_date, booked_time)
        except Exception:
            booked_timestamp = 0

        softdata_event = {
            "location": booked_origin,
            "timestamp": booked_timestamp,
            "event_description": "Softdata Upload",
            "status_external": "Softdata Upload",
            "status_internal": "softdata_upload"
        }

        tracking_sorted.append(softdata_event)

        return jsonify({
            "tracking_id": tracking_id,
            "trackHeader": track_header,
            "tracking": tracking_sorted
        }), 200

    except Exception as e:
        return jsonify({
            "tracking_id": tracking_id,
            "message": "Internal server error",
            "error": str(e)
        }), 500
    
@app.route('/api/offers', methods=['GET'])
def get_offers():
    try:
        page = int(request.args.get('page', 1))
        per_page = 5
        offset = (page - 1) * per_page

        mysql = get_mysql_connection()

        cursor = mysql.cursor()
        # Count total offers
        cursor.execute("SELECT COUNT(*) AS total_count FROM Offers")
        total_count = cursor.fetchone()['total_count']

        # Fetch paginated offers
        cursor.execute("""
            SELECT * FROM Offers
            ORDER BY creation_date DESC
            LIMIT %s OFFSET %s
        """, (per_page, offset))
        offers = cursor.fetchall()

        offer_list = []
        for offer in offers:
            offer_data = {
                "id": offer['id'],
                "title": offer['title'],
                "description": offer['description'],
                "image": offer['image'],
                "linkText": offer['linkText'],
                "linkIcon": offer['linkIcon'],
                "creation_date": offer['creation_date'].strftime("%Y-%m-%d %H:%M:%S")
            }
            offer_list.append(offer_data)

        return jsonify({
            "offers": offer_list,
            "total_pages": (total_count + per_page - 1) // per_page,
            "current_page": page,
            "has_next": page < (total_count + per_page - 1) // per_page,
            "has_prev": page > 1
        }), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500




@app.route('/api/notifications', methods=['GET'])
@token_required
def get_notifications(current_user):
    try:
        page = max(1, int(request.args.get('page', 1)))
        filter_type = request.args.get('filter', 'all')
        per_page = 10
        offset = (page - 1) * per_page

        raw_all = str(request.args.get('all_notifications', '0')).lower()
        all_notifications = raw_all in ('1', 'true', 'yes')

        allowed_filters = {'all', 'pending_confirm', 'confirm', 'dispatched', 'delivered'}
        if filter_type not in allowed_filters:
            filter_type = 'all'

        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            count_sql = "SELECT COUNT(*) AS total_count FROM Notifications WHERE notification_receiver = %s"
            count_args = [current_user]
            if filter_type != 'all':
                count_sql += " AND noti_type = %s"
                count_args.append(filter_type)
            cursor.execute(count_sql, tuple(count_args))
            total_count = cursor.fetchone()['total_count']

            fetch_sql = [
                "SELECT * FROM Notifications",
                "WHERE notification_receiver = %s"
            ]
            fetch_args = [current_user]

            if filter_type != 'all':
                fetch_sql.append("AND noti_type = %s")
                fetch_args.append(filter_type)

            if not all_notifications:
                fetch_sql.append("AND date_received < NOW()")

            fetch_sql.append("ORDER BY date_received DESC LIMIT %s OFFSET %s")
            fetch_args.extend([per_page, offset])

            cursor.execute(" ".join(fetch_sql), tuple(fetch_args))
            notifications = cursor.fetchall()

        def _to_iso_utc(dt):
            if not dt:
                return None
            return (dt.replace(tzinfo=timezone.utc) if dt.tzinfo is None else dt.astimezone(timezone.utc)).isoformat()

        notification_list = []
        for n in notifications:
            notification_list.append({
                "id": n['notification_id'],
                "sender": n['notification_sender'],
                "receiver": n['notification_receiver'],
                "message": n['notification_message'],
                "noti_type": n.get('noti_type'),
                "image": n.get('notification_image'),   # <-- only the filename (e.g., "delivered.png")
                "date_received": _to_iso_utc(n['date_received']),
                "read_status": _to_iso_utc(n['read_status']) if n.get('read_status') else None
            })

        total_pages = (total_count + per_page - 1) // per_page
        return jsonify({
            "notifications": notification_list,
            "total_pages": total_pages,
            "current_page": page,
            "has_next": page < total_pages,
            "has_prev": page > 1
        }), 200

    except Exception as e:
        return jsonify({"message": str(e)}), 500




def send_order_confirmation_notification(user_id, cart_id, cart_status):
    messages = {
        "pending_confirm": {
            "h1": "We're taking care of your order.",
            "p": "Your order's being checked, confirmation soon!",
            "image": "pending_confirm.png"
        },
        "confirm": {
            "h1": "Order confirmed.",
            "p": "Thanks for shopping with us—your order is being prepared with care. Pay online now and save ₹60 on delivery charges!",
            "image": "confirm.jpg"
        },
        "dispatched": {
            "h1": "Your order is dispatched.",
            "p": "Tracking will be shared soon.",
            "image": "dispatched.png"
        },
        "delivered": {
            "h1": "Please take a moment to give feedback and take 15 percent off on next order.",
            "p": "Thanks for choosing us. Please rate your experience and help us improve.",
            "image": "delivered.jpg"
        }
    }

    msg = messages.get(cart_status, {
        "h1": f"Your order #{cart_id} has been updated.",
        "p": "Please check the app for the latest updates on your order.",
        "image": "medingen-logo.jpg"
    })

    date_received = datetime.now(timezone.utc)
    notification_sender = "admin1"
    notification_receiver = user_id
    full_message = f"{msg['h1']} {msg['p']}".strip()

    mysql = get_mysql_connection()
    notification_triggered = False
    try:
        with mysql.cursor() as cursor:
            insert_sql = """
                INSERT INTO Notifications 
                (notification_sender, notification_receiver, notification_message, notification_image, date_received, read_status)
                VALUES (%s, %s, %s, %s, %s, %s)
            """
            cursor.execute(insert_sql, (
                notification_sender,
                notification_receiver,
                full_message,
                msg['image'],
                date_received,
                None
            ))
            mysql.commit()

            cursor.execute(
                "SELECT endpoint, p256dh, auth FROM PushSubscriptions WHERE user_id = %s",
                (notification_receiver,)
            )
            subscriptions = cursor.fetchall()

            if not subscriptions:
                print("Notification saved but no push subscriptions found for this user.")
                return False

            for subscription in subscriptions:
                api_data = {
                    "payload": {
                        "title": "Medingen",
                        "body": full_message,
                        "icon": "/android-chrome-192x192.png",
                        "target_url": "/notification"
                    },
                    "subscription": subscription,
                    "schedule_time": date_received.strftime("%Y-%m-%d %H:%M:%S")
                }

                api_url = "http://ec2-35-154-224-159.ap-south-1.compute.amazonaws.com/send_notification_247832438"
                try:
                    response = requests.post(api_url, json=api_data, timeout=5)
                    if response.status_code == 200:
                        notification_triggered = True
                    print(f"Push notification response: {response.status_code}")
                except Exception as e:
                    print(f"Error sending push notification: {str(e)}")

    except Exception as e:
        print(f"Error in send_order_confirmation_notification: {e}")
    finally:
        mysql.close()

    return notification_triggered


def send_welcome_default_notification(user_id):
        
    message = """
    📢 Exciting News from Medingen!
    🚀 The Medingen Pilot Study is NOW LIVE! 🚀

    Be among the first to experience high-quality generic medicines at unbeatable prices. For a limited time, enjoy a flat 20% discount on ALL products! 🏷️💊

    Don't miss out - order now and get your medicines delivered right to your doorstep! 🏠🚚

    T&Cs apply

    message 2 notification

    🎉 Special Offer Alert!

    🔍 Help Us Improve & Get Rewarded!

    During the Medingen Pilot Study, we’re offering unlimited cashback on your orders! 💸 Simply participate by identifying any errors or bugs in the MIG platform, and you'll earn cashback to use on your next purchase! 🛒

    Be a part of enhancing Medingen and enjoy exclusive rewards for your efforts!

    T&Cs apply
    """
    # after 30 mins of user creation
    date_received = datetime.now() + timedelta(minutes=3)
    notification_receiver = user_id
    
    mysql = get_mysql_connection()
    with mysql.cursor() as cursor:
        insert_sql = """
        INSERT INTO Notifications (notification_sender, notification_receiver, notification_message, date_received, read_status)
        VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(insert_sql, ("admin1", user_id, message, date_received, None))
        mysql.commit()
        
        cursor.execute("SELECT endpoint, p256dh, auth FROM PushSubscriptions WHERE user_id = %s", (notification_receiver,))
        subscriptions = cursor.fetchall()

        for subscription in subscriptions:
            api_data = {
                "payload": {
                    "title": "Medingen",
                    "body": message,
                    "icon": "/android-chrome-192x192.png",
                    "target_url": "/notification"
                },
                "subscription": subscription,
                "schedule_time": date_received.strftime("%Y-%m-%d %H:%M:%S")

            }

            api_url = f"http://ec2-35-154-224-159.ap-south-1.compute.amazonaws.com/send_notification_247832438"
            response = requests.post(api_url, json=api_data)
            print(f"Notification sent, status code: {response.status_code}")

    
@app.route('/api/banner', methods=['POST'])
def get_banner():
    data = request.json

    mysql = get_mysql_connection()
    with mysql.cursor() as cursor:
        query = "SELECT * FROM Banner WHERE section=%s"
        cursor.execute(query, data["section"])
        banners = cursor.fetchall()

    # Loop through each banner and parse the 'content' field
    banner_out = []
    for banner in banners:
        if 'content' in banner and isinstance(banner['content'], str):
            try:
                cleaned_content = banner['content'].replace('\r', '').replace('\n', '').replace('\t', '')
                c = json.loads(cleaned_content)
                c["tnc"] = banner.get("tnc", "")
                banner_out.append(c)  # Parse stringified JSON
            except json.JSONDecodeError:
                banner_out.append({})

    return jsonify(banner_out)


@app.route('/api/home_categories', methods=['GET'])
def get_home_categories():
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            query = """
                SELECT 
                    category_name,
                    display_name, 
                    show_on_home, 
                    category_image_url 
                FROM Category
                WHERE show_on_home > 0
                ORDER BY show_on_home ASC
            """
            cursor.execute(query)
            categories = cursor.fetchall()
    finally:
        mysql.close()

    return jsonify(categories)




def parse_json_field(value):
    if not value:
        return []
    try:
        if isinstance(value, str):
            return json.loads(value.strip())
        elif isinstance(value, list):
            return value
        else:
            return []
    except Exception as e:
        print("JSON parse error:", e, value)
        return []

@app.route('/api/product_details/<int:product_id>', methods=['GET'])
def get_product_details(product_id):
    try:
        product_name = request.args.get('name', None)
        show_hidden = request.args.get('show_hidden', False)

        mysql = get_mysql_connection()

        with mysql.cursor() as cursor:

            if not product_name:
                sql = """
                    SELECT 
                        p.*,
                        p.inventory_info_total_stock AS product_available,
                        c.*,
                        c.faq AS composition_faq
                    FROM Products p
                    LEFT JOIN composition_code c 
                        ON p.composition_code = c.composition_code
                    WHERE p.product_id = %s
                """
                if not show_hidden:
                    sql += " AND p.visibility_status = 'Published'"

                cursor.execute(sql, (product_id,))
                product = cursor.fetchone()

            else:
                sql = """
                    SELECT 
                        p.*,
                        p.inventory_info_total_stock AS product_available,
                        c.*,
                        c.faq AS composition_faq
                    FROM Products p
                    LEFT JOIN composition_code c 
                        ON p.composition_code = c.composition_code
                    WHERE p.product_name_url = %s
                """
                if not show_hidden:
                    sql += " AND p.visibility_status = 'Published'"

                cursor.execute(sql, (product_name,))
                product = cursor.fetchone()

            if not product:
                return jsonify({'message': 'Product not found'}), 404

            faq = parse_json_field(
                product.get("faq") or product.get("composition_faq")
            )

            recommended_results = []
            if product.get("rc") == 0:

                recommended_ids = product.get("recommended_product_id")

                try:
                    if recommended_ids is None:
                        recommended_ids = []

                    elif isinstance(recommended_ids, int):
                        recommended_ids = [recommended_ids]

                    elif isinstance(recommended_ids, str):
                        recommended_ids = [
                            int(x.strip()) for x in recommended_ids.split(",") if x.strip().isdigit()
                        ]

                    elif isinstance(recommended_ids, (list, tuple)):
                        recommended_ids = [
                            int(x) for x in recommended_ids if str(x).isdigit()
                        ]

                    else:
                        recommended_ids = []

                except Exception:
                    recommended_ids = []

                if isinstance(recommended_ids, list) and len(recommended_ids) > 0:
                    format_strings = ','.join(['%s'] * len(recommended_ids))

                    rec_query = f"""
                        SELECT
                            p.name AS product_name,
                            p.product_id,
                            p.photo AS images,
                            p.product_pricing_new,
                            p.product_pricing_old,
                            p.salt_name,
                            p.composition,
                            p.categories AS selected_category,
                            p.inStock,
                            p.rc,
                            p.manufacturer,
                            p.product_name_url,
                            c.category_outline_url
                        FROM Products p
                        LEFT JOIN (
                            SELECT DISTINCT category_name, category_outline_url
                            FROM Category
                        ) c ON p.categories = c.category_name
                        WHERE p.product_id IN ({format_strings})
                    """

                    if not show_hidden:
                        rec_query += " AND p.visibility_status = 'Published' "

                    rec_query += " AND p.product_pricing_new IS NOT NULL "
                    rec_query += " GROUP BY p.product_id ORDER BY p.product_pricing_new ASC"

                    cursor.execute(rec_query, recommended_ids)
                    recommended_results = cursor.fetchall()

            product_details = {
                "product_id": product.get("product_id"),
                "rackId": product.get("rack_id"),
                "totalStockQuantity": product.get("inventory_info_total_stock"),
                "departmentId": product.get("department_id"),
                "selectedCategory": product.get("categories"),
                "product_name_url": product.get("product_name_url"),

                "productPriceOld": str(product.get("product_pricing_old")),
                "productPriceNew": str(product.get("product_pricing_new")),
                "productCoupon": product.get("product_coupon_code"),

                "prescription_required": product.get("prescription_required"),
                "productName": product.get("name"),
                "saltName": product.get("salt_name"),

                "composition": product.get("composition"),
                "composition_code": product.get("composition_code"),
                "schedule_category": product.get("schedule_category"),

                "marketed_by": product.get("marketed_by"),
                "used_for": product.get("used_for"),
                "manufacturer": product.get("manufacturer"),

                "consumeType": product.get("consume_type"),
                "manufactureDate": str(product.get("manufacture_date")),
                "expiryDate": str(product.get("expiry")),

                "productDescription": product.get("product_description") or "",
                "descriptionLegacyUrl": product.get("description_url") or "",
                "pharmacistNote": product.get("pharmacist_note") or "",

                "benefits": parse_json_field(product.get("benefits")),
                "sideEffects": parse_json_field(product.get("side_effects")),
                "se_severity_options": parse_json_field(product.get("se_severity_options")),
                "se_type_options": parse_json_field(product.get("se_type_options")),

                "howToUse": parse_json_field(product.get("how_to_use")),
                "howItWorks": product.get("how_it_works") or "",

                "safetyAdvice": parse_json_field(product.get("safety_advice")),
                "drugInteractions": parse_json_field(product.get("drug_interactions")),
                "di_severity_options": parse_json_field(product.get("di_severity_options")),

                "faq": faq,
                "references": product.get("references") or "",

                "rc": product.get("rc"),
                "packaging": product.get("packaging"),

                "visibilityStatus": product.get("visibility_status"),
                "publishDate": str(product.get("publish_date")),
                "publishTime": str(product.get("publish_time")),

                "tags": product.get("tags").split(",") if product.get("tags") else [],
                "images": json.loads(product.get("photo", "[]")),

                "product_available": product.get("product_available", 0),
                "product_request": True,

                "rc_pharma_product_name": product.get("rc_pharma_product_name"),
                "inStock": bool(product.get("inStock")),

                "meta_keywords": product.get("meta_keywords"),
                "meta_description": product.get("meta_description"),
                "meta_title": product.get("meta_title"),

                "formulation": product.get("formulation"),

                "recommended_products": recommended_results
            }

            return jsonify(product_details)

    except Exception as e:
        return jsonify({"message": "An error occurred", "error": str(e)}), 500

    finally:
        try:
            cursor.close()
        except:
            pass




@app.route('/api/coupon_details/<string:couponcode>', methods=['GET'])
def get_coupon_details(couponcode):
    try:
        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            # Fetch product details
            sql = "SELECT * FROM Coupons WHERE coupon_code = %s"
            cursor.execute(sql, (couponcode,))
            coupon = cursor.fetchone()
            
            if not coupon:
                return jsonify({'message': 'Coupon not found'}), 404
            return jsonify(coupon)

    except Exception as e:
        return jsonify({'message': 'An error occurred', 'error': str(e)}), 500
    
    finally:
        cursor.close()

@app.route('/api/send_otp', methods=['POST'])
def send_otp():
    data = request.json
    phone_number = data.get('phone_number')
    otp = generate_random_digits(4)

    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            query = "SELECT * FROM Customers WHERE phonenumber = %s"
            cursor.execute(query, (phone_number,))
            user = cursor.fetchone()

            if user:
                query = "UPDATE Customers SET otp = %s WHERE phonenumber = %s"
                cursor.execute(query, (otp, phone_number))
            else:
                query = "INSERT INTO Customers (phonenumber, otp) VALUES (%s, %s)"
                cursor.execute(query, (phone_number, otp))

            mysql.commit()

        if not user or not user.get("otp_verification"):
            try:
                resp = create_notification_to_admin(
                    "system",
                    f"New customer with phone number {phone_number} has registered",
                    "/admin/customer-list/",
                    "new_customer"
                )
            except Exception as notif_exc:
                print("Notification error suppressed:", notif_exc)

        api_url = f"http://ec2-35-154-224-159.ap-south-1.compute.amazonaws.com/send_otpx3456sw?phone_number={phone_number}&otp={otp}"
        try:
            response = requests.get(api_url, timeout=8)
            if response.status_code == 200:
                return jsonify({"message": "OTP sent successfully"}), 200
            else:
                print(f'Error sending OTP API: {response.status_code} - {response.text}')
                return jsonify({"message": "Failed to send OTP"}), 500
        except requests.RequestException as req_exc:
            print("requests exception calling OTP API:", repr(req_exc))
            return jsonify({"message": "Failed to send OTP"}), 500

    except Exception as e:
        mysql.rollback()
        print(f"DB Error in send_otp: {e}", flush=True)
        return jsonify({"message": "An error occurred"}), 500

    finally:
        mysql.close()


@app.route('/api/update_order_auto_refill', methods=['POST'])
def update_order_auto_refill():
    data = request.json
    cart_id = data.get('cart_id')
    auto_refill = data.get('auto_refill')

    # Query database to verify credentials
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            query = "UPDATE Cart SET auto_refill = %s WHERE cart_id = %s"
            cursor.execute(query, (auto_refill, cart_id))
        
            mysql.commit()

            return jsonify({"message": "Auto refill status updated"}), 200

    except Exception as e:
        mysql.rollback()
        print(f"Error: {e}")
        return jsonify({"message": "An error occurred"}), 500

    finally:
        mysql.close()



@app.route('/api/check_customer', methods=['POST'])
def check_customer():
    data = request.json
    phone_number = data.get('phone_number')

    if not phone_number:
        return jsonify({"message": "Phone number is required"}), 400

    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            query = "SELECT otp_verification FROM Customers WHERE phonenumber = %s"
            cursor.execute(query, (phone_number,))
            result = cursor.fetchone()

            if result:
                otp_verified = result['otp_verification'] if isinstance(result, dict) else result[0]
                if otp_verified:
                    return jsonify({"exists": True, "message": "Customer found"}), 200
                else:
                    return jsonify({"exists": False, "message": "Customer not found"}), 200
            else:
                # Customer does not exist
                return jsonify({"exists": False, "message": "Customer not found"}), 200

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"message": "An error occurred"}), 500

    finally:
        mysql.close()


@app.route('/api/login_otp', methods=['POST'])
def login_otp():
    data = request.json
    phone_number = data.get('phone_number')
    otp = data.get('otp')  

    # Query database to verify credentials
    mysql = get_mysql_connection()
    with mysql.cursor() as cursor:
        query = "SELECT * FROM Customers WHERE phonenumber = %s AND otp = %s"
        cursor.execute(query, (phone_number, otp))
        user = cursor.fetchone()

        if user:
            # Update last_login_date and otp_verification
            update_query = """
                UPDATE Customers 
                SET last_login_date = %s, otp_verification = %s 
                WHERE customer_id = %s
            """
            cursor.execute(update_query, (datetime.now(), True, user['customer_id']))
            mysql.commit()
            token = jwt.encode({'customer_id': user['customer_id']}, app.config['SECRET_KEY'], algorithm="HS256")
            return jsonify({
                'token': token,
                'customer_id': user["customer_id"],
                'customer_name': user["customer_name"],
                'email': user["email"]
            }), 200
        else:
            return jsonify({'error': 'Invalid credentials'}), 401


@app.route('/api/create_password', methods=['POST'])
def create_password():
    data = request.json
    phone_number = data.get('phone_number')
    new_password = data.get('password')
    otp = data.get('otp')

    if not phone_number or not new_password or not otp:
        return jsonify({'error': 'Phone number, new password, and OTP are required'}), 400

    # Connect to the database
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            # Verify OTP and retrieve user details
            query = """
                SELECT customer_id, customer_name
                FROM Customers 
                WHERE phonenumber = %s AND otp = %s
            """
            cursor.execute(query, (phone_number, otp))
            user = cursor.fetchone()

            if not user:
                return jsonify({'error': 'Phone number or OTP invalid'}), 404

            # Update the password
            update_query = "UPDATE Customers SET password = %s WHERE customer_id = %s"
            cursor.execute(update_query, (new_password, user['customer_id']))
            mysql.commit()

            return jsonify({'message': 'Password updated successfully', "customer_name": user["customer_name"]}), 200

    except Exception as e:
        print(e)
        return jsonify({'error': str(e)}), 500

    finally:
        # Ensure the database connection is properly closed
        mysql.close()


@app.route('/api/login_password', methods=['POST'])
def login_password():
    data = request.json
    phone_number = data.get('phone_number')
    password = data.get('password')

    if not phone_number or not password:
        return jsonify({'error': 'Phone number and password are required'}), 400

    # Query database to verify credentials
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            query = "SELECT * FROM Customers WHERE phonenumber = %s AND password = %s"
            cursor.execute(query, (phone_number,password))
            user = cursor.fetchone()

            if user:
                # Update last login date
                update_query = "UPDATE Customers SET last_login_date = %s WHERE customer_id = %s"
                cursor.execute(update_query, (datetime.now(), user['customer_id']))
                mysql.commit()

                # Generate JWT token
                token = jwt.encode({'customer_id': user['customer_id']}, app.config['SECRET_KEY'], algorithm="HS256")
                
                return jsonify({
                    'token': token,
                    'customer_id': user['customer_id'],
                    'customer_name': user['customer_name'],
                    'email': user['email']
                }), 200
            else:
                return jsonify({'error': 'Invalid credentials'}), 401

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': 'An error occurred'}), 500


@app.route('/api/update_profile', methods=['POST'])
@token_required
def update_profile(current_user):
    data = request.json

    if not data:
        return jsonify({"message": "No data provided"}), 400

    customer_id = current_user
    if not customer_id:
        return jsonify({"message": "customer_id is required"}), 400

    # Build the dynamic SQL UPDATE statement
    update_fields = []
    update_values = []
    for field in [
        'customer_name', 'email', 'billing_address', 'pincode',
        'state', 'city', 'dob', 'profile_picture', "phonenumber",
        'gender', 'blood_group', 'password'
    ]:
        if field in data:
            update_fields.append(f"{field} = %s")
            update_values.append(data[field])

    if not update_fields:
        return jsonify({"message": "No fields to update"}), 400

    update_values.append(customer_id)  # Append the customer_id for the WHERE clause
    update_query = f"""
    UPDATE Customers
    SET {', '.join(update_fields)}
    WHERE customer_id = %s
    """

    try:
        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            cursor.execute(update_query, update_values)
            mysql.commit()


        return jsonify({"message": "Profile updated successfully"}), 200
    except Exception as e:
        print(e)
        return jsonify({"message": "An error occurred while updating the profile"}), 500



@app.route('/api/update_prescription', methods=['POST'])
@token_required
def update_prescription(customer_id):
    data = request.json
    if not data:
        return jsonify({"message": "No data provided"}), 400

    prescription_image_url = data.get('prescription_image_url')
    prescription_date = data.get('prescription_date')
    prescription_status = data.get('prescription_status')
    prescription_comments = data.get('prescription_comments', '')
    prescription_name = data.get("prescription_name", "")

    if not all([prescription_image_url, prescription_date]):
        return jsonify({"message": "Missing required fields"}), 400

    insert_query = """
    INSERT INTO Prescription (prescription_name, prescription_image_url, prescription_date, prescription_status, prescription_comments, customer_id)
    VALUES (%s, %s, %s, %s, %s, %s)
    """

    try:
        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            cursor.execute(insert_query, (prescription_name, prescription_image_url, prescription_date, prescription_status, prescription_comments, customer_id))
            mysql.commit()
            last_inserted_id = cursor.lastrowid


        return jsonify({"message": "Prescription inserted successfully", "prescription_id": last_inserted_id}), 200
    except Exception as e:
        print(e)
        return jsonify({"message": "An error occurred while inserting the prescription"}), 500


@app.route('/api/place_prescription', methods=['POST'])
@token_required
def place_prescription(customer_id):
    data = request.json
    if not data:
        return jsonify({"message": "No data provided"}), 400

    prescription_id = data.get("prescription_id", "")

    try:
        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            
            insert_query = """
            UPDATE Prescription SET prescription_status='ORDER_PLACED' WHERE prescription_id=%s
            """

            cursor.execute(insert_query, (prescription_id,))
            mysql.commit()

            product_id = 38743
            quantity = 1

            # Check if the customer has an existing cart
            query = "SELECT cart_id FROM Cart WHERE customer_id = %s AND cart_status IN ('active', 'pending_confirm', 'confirm')"
            cursor.execute(query, (customer_id,))
            cart = cursor.fetchone()
            if cart:
                cart_id = cart['cart_id']
                # Update existing cart
                query = "UPDATE Cart SET product_ids_qty = CONCAT(IFNULL(product_ids_qty, ''), %s, ':', %s, ';'), cart_status = 'pending_confirm', prescription_id = %s WHERE cart_id = %s"
                cursor.execute(query, (product_id, quantity, prescription_id, cart_id))
            else:
                # Create new cart
                query = "INSERT INTO Cart (customer_id, cart_status, product_ids_qty, prescription_id) VALUES (%s, 'pending_confirm', %s, %s)"
                cursor.execute(query, (customer_id, f"{product_id}:{quantity};", prescription_id))
                cart_id = cursor.lastrowid

            mysql.commit()

        create_notification_to_admin(customer_id, f"Cart updated for cart {cart_id}", f"/admin/customer_detail/{customer_id}", "order")

        return jsonify({"message": "Prescription placed successfully"}), 200
    except Exception as e:
        print(e)
        return jsonify({"message": "An error occurred while placing the prescription"}), 500


@app.route('/api/list_prescriptions', methods=['GET', "POST"])
@token_required
def list_prescriptions(customer_id):

    select_query = """
    SELECT prescription_name, prescription_id, prescription_image_url, prescription_date, prescription_status, prescription_comments, last_used_date, associated_products
    FROM Prescription
    WHERE customer_id = %s
    """

    try:
        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            cursor.execute(select_query, (customer_id,))
            prescriptions = cursor.fetchall()

            for i, prescription in enumerate(prescriptions):
                prescription["associated_products"] = prescription["associated_products"].split(",")
                # Fetch associated products
                if prescription["associated_products"]:
                    product_ids = prescription["associated_products"]
                    for j, product_id in enumerate(product_ids):
                        cursor.execute("SELECT name FROM Products WHERE product_id = %s", (product_id,))
                        product = cursor.fetchone()
                        if product:
                            product_name = product["name"]
                            prescriptions[i]["associated_products"][j] = {"product_id": product_id, "product_name": product_name}

        return jsonify({"prescriptions": prescriptions}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"message": "An error occurred while retrieving prescriptions"}), 500

@app.route('/api/get_profile', methods=['GET'])
@token_required
def get_profile(current_user):
    customer_id = current_user

    if not customer_id:
        return jsonify({"message": "customer_id is required"}), 400

    select_query = """
    SELECT
        customer_name AS name,
        email,
        dob,
        phonenumber AS phone,
        state,
        total_mig_coins_available AS coins,
        'light' AS theme,
        profile_picture AS profilePicture,
        customer_type,
        customer_category,
        gender,
        blood_group
    FROM Customers
    WHERE customer_id = %s
    """

    try:
        mysql = get_mysql_connection()
        with mysql.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(select_query, (customer_id,))
            result = cursor.fetchone()

        if result:
            return jsonify(result), 200
        else:
            return jsonify({"message": "Profile not found"}), 404

    except Exception:
        return jsonify({"message": "An error occurred while retrieving the profile"}), 500

@app.route('/api/update_address', methods=['POST'])
@token_required
def update_address(current_user):
    data = request.json
    if not data:
        return jsonify({"message": "No data provided"}), 400

    customer_id = current_user
    address_id = data.get('id')
    if not address_id:
        return jsonify({"message": "Address ID is required"}), 400

    new_address = data.get('address')
    if not new_address:
        return jsonify({"message": "Address data is required"}), 400

    try:
        mysql = get_mysql_connection()
        if not mysql:
            return jsonify({"message": "Database connection failed"}), 500
        
        cursor = mysql.cursor()

        update_query = """
        UPDATE Addresses
        SET 
            address1 = %s,
            name = %s,
            pincode = %s,
            state = %s,
            type = %s,
            `default` = %s,
            phone_number = %s,
            alt_phone_number = %s
        WHERE id = %s AND customer_id = %s
        """

        cursor.execute(update_query, (
            new_address.get('address1'),
            new_address.get('name'),
            new_address.get('pincode'),
            new_address.get('state'),
            new_address.get('type'),
            new_address.get('default'),
            new_address.get('phone_number', ""),
            new_address.get('alt_phone_number', ""),  
            address_id,
            customer_id
        ))

        mysql.commit()
        cursor.close()
        mysql.close()

        return jsonify({"message": "Address updated successfully"}), 200

    except Exception as e:
        print(e)
        return jsonify({"message": "An error occurred while updating the address"}), 500


@app.route('/api/add_address', methods=['POST'])
@token_required
def add_address(current_user):
    data = request.json
    if not data:
        return jsonify({"message": "No data provided"}), 400

    new_address = data.get('address')
    if not new_address:
        return jsonify({"message": "Address data is required"}), 400

    try:
        mysql = get_mysql_connection()
        if not mysql:
            return jsonify({"message": "Database connection failed"}), 500

        cursor = mysql.cursor()

        # Insert new address into the Addresses table
        insert_query = """
        INSERT INTO Addresses (customer_id, address1, name, pincode, state, type, `default`, phone_number)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(insert_query, (
            current_user,
            new_address.get('address1'),
            new_address.get('name'),
            new_address.get('pincode'),
            new_address.get('state'),
            new_address.get('type'),
            False,  # New address is default,
            new_address.get('phone_number', "")
        ))
        mysql.commit()
        cursor.close()
        mysql.close()

        return jsonify({"message": "Address added successfully"}), 200
    except Exception as e:
        print(e)
        return jsonify({"message": "An error occurred while adding the address"}), 500
    
@app.route('/api/delete_address', methods=['POST'])
def admin_delete_address():
    # Read from query parameters
    address_id = request.args.get('id')
    customer_id = request.args.get('customer')

    if not address_id:
        return jsonify({"message": "Address ID is required"}), 400

    if not customer_id:
        return jsonify({"message": "Customer ID is required"}), 400

    try:
        mysql = get_mysql_connection()
        if not mysql:
            return jsonify({"message": "Database connection failed"}), 500

        cursor = mysql.cursor()

        # Delete address from the Addresses table
        delete_query = """
        DELETE FROM Addresses
        WHERE id = %s AND customer_id = %s
        """
        cursor.execute(delete_query, (address_id, customer_id))
        mysql.commit()
        cursor.close()
        mysql.close()

        return jsonify({"message": "Address deleted successfully"}), 200
    except Exception as e:
        print(e)
        return jsonify({"message": "An error occurred while deleting the address"}), 500


@app.route('/api/delete_prescription', methods=['POST'])
@token_required
def delete_prescription(current_user):
    data = request.json
    if not data:
        return jsonify({"message": "No data provided"}), 400

    prescription_id = data.get('id')
    if not prescription_id:
        return jsonify({"message": "Prescription ID is required"}), 400

    try:
        mysql = get_mysql_connection()
        if not mysql:
            return jsonify({"message": "Database connection failed"}), 500

        cursor = mysql.cursor()

        # Delete prescription from the Prescriptions table
        delete_query = """
        DELETE FROM Prescription
        WHERE prescription_id = %s AND customer_id = %s
        """
        print(delete_query, prescription_id, current_user)
        cursor.execute(delete_query, (prescription_id, current_user))
        mysql.commit()
        cursor.close()
        mysql.close()

        return jsonify({"message": "Prescription deleted successfully"}), 200
    except Exception as e:
        print(e)
        return jsonify({"message": "An error occurred while deleting the prescription"}), 500

@app.route('/api/list_addresses', methods=['GET'])
@token_required
def list_addresses(current_user):
    print("list address", current_user)
    try:
        mysql = get_mysql_connection()
        if not mysql:
            return jsonify({"message": "Database connection failed"}), 500

        cursor = mysql.cursor()

        select_query = """
        SELECT 
            id, 
            address1, 
            name, 
            pincode, 
            state, 
            type, 
            `default`, 
            phone_number,
            alt_phone_number
        FROM Addresses
        WHERE customer_id = %s
        """
        
        cursor.execute(select_query, (current_user,))
        result = cursor.fetchall()

        cursor.close()
        mysql.close()

        if result:
            addresses = [dict(row) for row in result]
            return jsonify({"addresses": addresses}), 200
        else:
            return jsonify({"addresses": [], "message": "No addresses found"}), 200

    except Exception as e:
        print(e)
        return jsonify({"message": "An error occurred while retrieving the addresses"}), 500


@app.route('/api/select_address', methods=['POST'])
@token_required
def select_address(current_user):
    data = request.json
    if not data:
        return jsonify({"message": "No data provided"}), 400

    address_id = data.get('id')
    if not address_id:
        return jsonify({"message": "Address ID is required"}), 400

    try:
        mysql = get_mysql_connection()
        if not mysql:
            return jsonify({"message": "Database connection failed"}), 500

        cursor = mysql.cursor()

        # Set the selected address as default and others as non-default
        update_query = """
        UPDATE Addresses
        SET `default` = CASE WHEN id = %s THEN TRUE ELSE FALSE END
        WHERE customer_id = %s
        """
        cursor.execute(update_query, (address_id, current_user))
        mysql.commit()
        cursor.close()
        mysql.close()

        return jsonify({"message": "Default address updated successfully"}), 200
    except Exception as e:
        print(e)
        return jsonify({"message": "An error occurred while updating the address"}), 500



@app.route('/api/get_default_address', methods=['GET'])
@token_required
def get_default_address(current_user):
    try:
        mysql = get_mysql_connection()
        if not mysql:
            return jsonify({"message": "Database connection failed"}), 500

        cursor = mysql.cursor()

        select_query = """
        SELECT * FROM Addresses
        WHERE customer_id = %s AND `default` = TRUE
        """
        cursor.execute(select_query, (current_user,))
        result = cursor.fetchone()
        cursor.close()
        mysql.close()

        if result:
            return jsonify({"default_address": result}), 200
        else:
            return jsonify({"message": "No default address found"}), 404
    except Exception as e:
        print(e)
        return jsonify({"message": "An error occurred while retrieving the default address"}), 500

@app.route('/api/rewards', methods=['GET'])
@token_required
def get_rewards(customer_id):
    try:
        page = int(request.args.get('page', 1))
        per_page = 10
        offset = (page - 1) * per_page

        mysql = get_mysql_connection()

        cursor = mysql.cursor()
        # Count total transactions
        cursor.execute("""
            SELECT COUNT(*) AS total_count 
            FROM RewardsTransaction
            WHERE customer_id = %s
        """, (customer_id,))
        total_count = cursor.fetchone()['total_count']

        # Fetch paginated transactions
        cursor.execute("""
            SELECT
                id,
                description,
                orderid,
                CONCAT(date, ' ', time) AS date_time,
                reward,
                iconType,
                amount,
                percentage,
                expires
            FROM RewardsTransaction
            WHERE customer_id = %s
            ORDER BY date DESC, time DESC
            LIMIT %s OFFSET %s
        """, (customer_id, per_page, offset))
        transactions = cursor.fetchall()

        transaction_list = []
        for transaction in transactions:
            # Parse the combined date_time field
            date_time = datetime.strptime(transaction['date_time'], "%Y-%m-%d %H:%M:%S")
            
            transaction_data = {
                "id": transaction['id'],
                "description": transaction['description'],
                "orderid": transaction['orderid'],
                "date": date_time.strftime("%Y-%m-%d"),
                "time": date_time.strftime("%H:%M:%S"),
                "reward": transaction['reward'],
                "iconType": transaction['iconType'],
                "amount": transaction['amount'],
                "percentage": transaction['percentage'],
                "expires": transaction['expires'].strftime("%Y-%m-%d")  # Assuming expires is a date
            }
            transaction_list.append(transaction_data)

        return jsonify({
            "customer_id": customer_id,
            "transactions": transaction_list,
            "total_pages": (total_count + per_page - 1) // per_page,
            "current_page": page,
            "has_next": page < (total_count + per_page - 1) // per_page,
            "has_prev": page > 1
        }), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500

@app.route('/api/rewards-summary', methods=['GET'])
@token_required
def get_rewards_summary(customer_id):
    print(customer_id)
    try:
        mysql = get_mysql_connection()
        cursor = mysql.cursor()

        # Define the date one month from now
        one_month_from_now = datetime.now() + timedelta(days=30)
        formatted_one_month_from_now = one_month_from_now.strftime("%Y-%m-%d")

        # Fetch total rewards for the customer
        cursor.execute("""
            SELECT SUM(reward) AS total_amount 
            FROM RewardsTransaction 
            WHERE iconType = 'primary' AND customer_id = %s
        """, (customer_id,))
        total_rewards = cursor.fetchone()['total_amount'] or 0
        print(total_rewards)

        # Fetch total redemptions for the customer
        cursor.execute("""
            SELECT SUM(reward) AS total_amount 
            FROM RewardsTransaction 
            WHERE iconType = 'secondary' AND customer_id = %s
        """, (customer_id,))
        total_redemptions = cursor.fetchone()['total_amount']
        if not total_redemptions:
            total_redemptions = 0
        total_redemptions = -total_redemptions
        

        # Fetch total expiring rewards for the customer
        cursor.execute("""
            SELECT SUM(reward) AS total_amount 
            FROM RewardsTransaction 
            WHERE iconType = 'primary' AND customer_id = %s AND expires <= %s AND redeemed = 0
        """, (customer_id, formatted_one_month_from_now))
        expiring_rewards = cursor.fetchone()['total_amount'] or 0

        return jsonify({
            "available": "{:.2f}".format(total_rewards - total_redemptions),
            "overall": "{:.2f}".format(total_rewards),
            "expiring": "{:.2f}".format(expiring_rewards)
        }), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500


# Add Reward API
@app.route("/api/complete_mig_payment", methods=["POST"])
@token_required
def complete_mig_payment(customer_id):
    try:
        data = request.json
        cart_id = data.get("cart_id")
        total_amount = data.get("total_amount")
        coupon_code_used = data.get("coupon_code_used")

       
        connection = get_mysql_connection()
        cursor = connection.cursor()



        connection.commit()
        cursor.close()
        connection.close()

        return jsonify({"message": "MIG payment done!"}), 201
    except Exception as e:
        return jsonify({"message": str(e)}), 500



@app.route('/api/all_categories', methods=['GET'])
def get_all_categories():
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            query = """
                SELECT category_id, category_name, category_description, category_image_url, category_outline_url,
                       category_datetime_created, display_name, show_on_home, main_category
                FROM Category
                ORDER BY category_datetime_created DESC
            """
            cursor.execute(query)
            categories = cursor.fetchall()
        return jsonify(categories), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()


@app.route('/api/all_blog_categories', methods=['GET'])
def get_all_blog_categories():
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            query = """
                SELECT id, category_name, category_display_name, category_image_url, 
                       category_description, category_on_home
                FROM BlogCategories
                ORDER BY id DESC
            """
            cursor.execute(query)
            categories = cursor.fetchall()
        return jsonify(categories), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()
        

@app.route('/api/get_salt', methods=['POST'])
def get_salt():
    data = request.get_json()
    salt_name = data.get('salt_name')
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            query = """
                SELECT id, composition_code, composition, description_url, rc, packaging
                FROM composition_code
                WHERE composition = %s
                ORDER BY id DESC
            """
            cursor.execute(query, (salt_name,))
            salt = cursor.fetchone()
        return jsonify(salt), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()


@app.route('/api/search_composition_code', methods=['POST'])
def search_composition_code():
    search_term = request.json.get('search_term')
    if not search_term:
        return jsonify({"error": "Search term is required"}), 400

    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            query = """
                SELECT 
                    id, 
                    composition_code, 
                    composition, 
                    description_url, 
                    rc, 
                    packaging,
                    LOCATE(%s, composition) AS position,
                    LENGTH(composition) AS length
                FROM composition_code
                WHERE 
                    composition LIKE %s
                ORDER BY length ASC, position DESC,  id DESC
                LIMIT 5
            """
            wildcard_term = f"%{search_term}%"
            cursor.execute(query, (search_term, wildcard_term))
            composition_codes = cursor.fetchall()

        return jsonify(composition_codes), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()



@app.route("/sitemap.html", methods=["GET"])
def render_sitemap():
    static_routes = [
        "https://medingen.in/",
        "https://medingen.in/dashboard",
        "https://medingen.in/offers",
        "https://medingen.in/rewards",
        "https://medingen.in/profile",
        "https://medingen.in/orders",
        "https://medingen.in/privacy-policy",
        "https://medingen.in/notification",
        "https://medingen.in/terms-and-conditions",
        "https://medingen.in/reminder",
        "https://medingen.in/cart",
        "https://medingen.in/blogs",
        "https://medingen.in/about"
    ]
    
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            # Fetch blog URLs
            cursor.execute("SELECT blog_name, blog_url FROM Blogs WHERE 1=1 AND visibility_status = 'Published' ORDER BY blog_rc_priority ASC")
            blog_routes = [(row['blog_name'], f"https://medingen.in/blogs/{row['blog_url']}") for row in cursor.fetchall()]
            
            # Fetch product URLs
            cursor.execute("SELECT categories, product_name_url FROM Products")
            product_routes = [f"https://medingen.in/product/{row['product_name_url']}" for row in cursor.fetchall()]
            
            # Fetch composition URLs
            cursor.execute("SELECT composition FROM composition_code")
            composition_routes = [f"https://medingen.in/salt/{row['composition']}" for row in cursor.fetchall()]
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()

    # Organize URLs alphabetically for Salt and Product pages
    salt_routes = sorted(composition_routes)
    product_routes = sorted(product_routes)

    # Helper function to group routes by their initial alphabet letter
    def group_by_letter(routes, index=4):
        grouped = {}
        for route in routes:
            try:
                first_letter = route.split("/")[index][0].upper()  # Get the first letter of the last part of URL
                if first_letter not in grouped:
                    grouped[first_letter] = []
                grouped[first_letter].append(route)
            except IndexError:
                continue  # Skip the URLs that don't have enough segments
        return grouped

    salt_grouped = group_by_letter(salt_routes)
    product_grouped = group_by_letter(product_routes, 4)

    # HTML content for the sitemap page
    html_content = """
    <html>
    <head><title>Sitemap</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f4f4f4;
        }
        h1 {
            color: #333;
        }
        h2, h3 {
            color: #0056b3;
        }
        .container {
            display: grid;
            grid-template-columns: 1fr 1fr 1fr;
            gap: 20px;
            max-width: 100%;
            padding: 10px;
        }
        .table {
            display: grid;
            grid-template-columns: 1fr 1fr 1fr;
            gap: 10px;
            border: 1px solid #ddd;
            padding: 10px;
            background-color: #fff;
            box-shadow: 0 2px 5px rgba(0, 0, 0, 0.1);
        }
        .table div {
            padding: 10px 0;
            text-align: center;
            word-wrap: break-word;
            border: 1px solid #ddd;
        }
        a {
            color: #0066cc;
            text-decoration: none;
            word-wrap: break-word;
        }
        a:hover {
            text-decoration: underline;
        }
        @media screen and (max-width: 768px) {
            .container {
                grid-template-columns: 1fr 1fr;
            }
            .table {
                grid-template-columns: 1fr;
            }
            .table div {
                font-size: 14px;
                padding: 8px;
            }
            h2, h3 {
                font-size: 16px;
            }
        }
    </style>
    </head>
    <body>
    <h1>Sitemap for medingen.in</h1>
    
    <h2>General Pages</h2>
    <div class="container">
    """
    for url in static_routes:
        try:
            html_content += f"  <div><a href='{url}'>{url}</a></div>\n"
        except IndexError:
            continue  # Skip if the index doesn't exist
    html_content += "</div>\n"

    html_content += "<h2>Blog Pages</h2>\n"
    for url in blog_routes:
        try:
            html_content += f"  <div><a href='{url[1]}'>{url[0]}</a></div>\n"
        except IndexError:
            continue  # Skip if the index doesn't exist
    html_content += "</div>\n"

    
    # Salt pages section (A-Z)
    html_content += "<h2>Salt Pages (A-Z)</h2>\n"
    for letter in sorted(salt_grouped.keys()):
        html_content += f"<h3>{letter}</h3>\n"
        html_content += "<div class='table'>\n"
        for url in salt_grouped[letter]:
            try:
                url_part = url.split("/")[4]  # Extracting the relevant part
                html_content += f"    <div><a href='{url}'>{url_part}</a></div>\n"
            except IndexError:
                continue  # Skip if the index doesn't exist
        html_content += "</div>\n"

    # Product pages section (A-Z)
    html_content += "<h2>Product Pages (A-Z)</h2>\n"
    for letter in sorted(product_grouped.keys()):
        html_content += f"<h3>{letter}</h3>\n"
        html_content += "<div class='table'>\n"
        for url in product_grouped[letter]:
            try:
                url_part = url.split("/")[4]  # Extracting the relevant part
                html_content += f"    <div><a href='{url}'>{url_part}</a></div>\n"
            except IndexError:
                continue  # Skip if the index doesn't exist
        html_content += "</div>\n"
    
    html_content += "</body></html>"

    return Response(html_content, mimetype="text/html")





@app.route("/sitemap.xml", methods=["GET"])
def generate_sitemap():
    # Fixed timestamp: today's date at 00:00:00 UTC
    now = datetime.utcnow().date().isoformat() + "T00:00:00+00:00"

    static_routes = [
        "https://medingen.in/",
        "https://medingen.in/dashboard",
        "https://medingen.in/offers",
        "https://medingen.in/rewards",
        "https://medingen.in/profile",
        "https://medingen.in/orders",
        "https://medingen.in/privacy-policy",
        "https://medingen.in/notification",
        "https://medingen.in/terms-and-conditions",
        "https://medingen.in/reminder",
        "https://medingen.in/cart",
        "https://medingen.in/blogs",
        "https://medingen.in/about"
    ]

    blog_routes = []
    product_routes = []
    composition_routes = []

    try:
        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            cursor.execute("""
                SELECT blog_url FROM Blogs 
                WHERE visibility_status = 'Published'
                ORDER BY blog_rc_priority ASC
            """)
            blog_routes = [
                f"https://medingen.in/blogs/{row['blog_url']}"
                for row in cursor.fetchall()
            ]

            cursor.execute("SELECT product_name_url FROM Products")
            product_routes = [
                f"https://medingen.in/product/{row['product_name_url']}"
                for row in cursor.fetchall()
            ]

            cursor.execute("SELECT composition FROM composition_code")
            composition_routes = [
                f"https://medingen.in/salt/{row['composition']}"
                for row in cursor.fetchall()
            ]
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()

    # Start building the sitemap index XML
    xml_parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<sitemapindex xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">'
    ]

    def sitemap_block(urls):
        block = []
        for url in urls:
            block.append("  <sitemap>")
            block.append(f"    <loc>{escape(url)}</loc>")
            block.append(f"    <lastmod>{now}</lastmod>")
            block.append("  </sitemap>")
        return block

    xml_parts += sitemap_block(static_routes)
    xml_parts += sitemap_block(blog_routes)
    xml_parts += sitemap_block(product_routes)
    xml_parts += sitemap_block(composition_routes)

    xml_parts.append('</sitemapindex>')
    xml_parts.append('<!-- XML Sitemap generated by Medingen -->')

    return Response("\n".join(xml_parts), mimetype="application/xml")



def _load_json_field(row, field):
    v = row.get(field)
    if not v:
        return []
    try:
        if isinstance(v, (list, dict)):
            return v
        return json.loads(v)
    except Exception:
        return []

@app.route('/api/all_blogs', methods=['GET'])
def get_all_blogs():
    mysql = get_mysql_connection()
    popular = request.args.get('popular', False)
    category_id = request.args.get('category_id', None)
    show_hidden = request.args.get('show_hidden', False)

    if popular:
        order_col = "blog_visit_count"
        order_dir = "DESC"
    else:
        order_col = "blog_rc_priority"
        order_dir = "ASC"
    
    try:
        with mysql.cursor() as cursor:
            # Base query
            query = f"""
                SELECT id, blog_name, blog_image_url, blog_category_id, blog_description_url, blog_visit_count, 
                       blog_rc_priority, blog_created_date, blog_url, meta_keywords, meta_description, meta_title, 
                       meta_author, meta_author_title, meta_author_profile_url, visibility_status,
                       blog_likes_count, blog_comments_count
                FROM Blogs WHERE 1=1
            """
            
            if category_id:
                query += " AND blog_category_id = %s "
            
            if not show_hidden:
                query += " AND visibility_status = 'Published' "
            
            query += f" ORDER BY {order_col} {order_dir}"

            if category_id:
                cursor.execute(query, (category_id,))
            else:
                cursor.execute(query)

            blogs = cursor.fetchall()

            # Now, include likes, comments counts, and meta data for each blog
            blog_data = []
            for blog in blogs:
                blog_id = blog['id']
                blog_url = blog['blog_url']

                # Get likes and comments count directly from the columns
                likes_count = blog['blog_likes_count'] if blog['blog_likes_count'] is not None else 0
                comments_count = blog['blog_comments_count'] if blog['blog_comments_count'] is not None else 0

                # Get the likes and liked_by_user status from the JSON field (if applicable)
                cursor.execute("SELECT blog_likes FROM Blogs WHERE blog_url = %s LIMIT 1", (blog_url,))
                row = cursor.fetchone()
                likes = _load_json_field(row or {}, "blog_likes") if row else []
                liked_by_user = False
                auth = request.headers.get("Authorization", "")
                if auth.startswith("Bearer ") and app and jwt:
                    token = auth.split(" ", 1)[1]
                    try:
                        data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
                        user_id = data.get("customer_id") or data.get("user")
                        if user_id:
                            liked_by_user = int(user_id) in [int(x) for x in likes] if likes else False
                    except Exception:
                        liked_by_user = False

                # Add meta data, likes, and comments counts to the blog data
                blog_dict = dict(blog)
                blog_dict['comments_count'] = comments_count
                blog_dict['likes_count'] = likes_count
                blog_dict['liked_by_user'] = liked_by_user
                blog_data.append(blog_dict)

        return jsonify(blog_data), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()


# Updated get_blog with likes and comments
@app.route('/api/get_blog', methods=['POST'])
def get_blog():
    blog_name = request.json.get('blog_url')
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            query = """
                SELECT id, blog_name, blog_image_url, blog_category_id, blog_description_url, blog_visit_count, blog_rc_priority, blog_created_date, meta_keywords, blog_url, meta_description, meta_title, meta_author, meta_author_title, meta_author_profile_url, visibility_status
                FROM Blogs
                WHERE blog_url = %s
                ORDER BY id DESC
            """
            cursor.execute(query, (blog_name,))
            blog = cursor.fetchone()

            if blog:
                blog_url = blog['blog_url']

                # Get comments
                cursor.execute("SELECT blog_comments FROM Blogs WHERE blog_url = %s LIMIT 1", (blog_url,))
                row = cursor.fetchone()
                comments = _load_json_field(row, "blog_comments") if row else []

                # Get likes count and user like status
                cursor.execute("SELECT blog_likes, blog_likes_count FROM Blogs WHERE blog_url = %s LIMIT 1", (blog_url,))
                row = cursor.fetchone()
                likes = _load_json_field(row or {}, "blog_likes") if row else []
                likes_count = int(row.get("blog_likes_count") or len(likes)) if row else 0
                liked_by_user = False
                auth = request.headers.get("Authorization", "")
                if auth.startswith("Bearer ") and app and jwt:
                    token = auth.split(" ", 1)[1]
                    try:
                        data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
                        user_id = data.get("customer_id") or data.get("user")
                        if user_id:
                            liked_by_user = int(user_id) in [int(x) for x in likes] if likes else False
                    except Exception:
                        liked_by_user = False

                # Add likes and comments to the blog data
                blog['comments'] = comments
                blog['likes_count'] = likes_count
                blog['liked_by_user'] = liked_by_user

        return jsonify(blog), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()


@app.route("/api/blogs/<string:blog_url>/comments", methods=["POST"])
@token_required
def post_comment(current_user, blog_url):
    """
    Expects JSON: { "comment_text": "...", "parent_comment_id": <int|null> }
    current_user is injected by token_required (customer_id)
    Comments are stored in Blogs.blog_comments as JSON array of comment objects:
      { comment_id, customer_id, customer_name, profile_picture, comment_text, parent_comment_id, created_at }
    """
    try:
        data = request.get_json() or {}
        text = (data.get("comment_text") or "").strip()
        parent = data.get("parent_comment_id")
        if not text:
            return jsonify({"error": "comment_text is required"}), 400

        conn = get_mysql_connection()
        with conn.cursor() as cursor:
            cursor.execute("SELECT blog_comments FROM Blogs WHERE blog_url = %s LIMIT 1", (blog_url,))
            row = cursor.fetchone()
            existing = _load_json_field(row or {}, "blog_comments")

            # generate new comment_id (use max+1)
            max_id = 0
            for c in existing:
                try:
                    if int(c.get("comment_id", 0)) > max_id:
                        max_id = int(c.get("comment_id", 0))
                except Exception:
                    continue
            new_id = max_id + 1

            # Try to get customer name/profile (if Customers exists)
            customer_name = None
            profile_picture = None
            try:
                cursor.execute("SELECT customer_name, profile_picture FROM Customers WHERE customer_id = %s LIMIT 1", (current_user,))
                cu = cursor.fetchone()
                if cu:
                    customer_name = cu.get("customer_name")
                    profile_picture = cu.get("profile_picture")
            except Exception:
                # ignore if Customers table absent
                pass

            new_comment = {
                "comment_id": new_id,
                "customer_id": current_user,
                "customer_name": customer_name,
                "profile_picture": profile_picture,
                "comment_text": text,
                "parent_comment_id": parent,
                "created_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            }
            existing.append(new_comment)
            # update Blogs.blog_comments JSON and increment blog_comments_count
            comments_json = json.dumps(existing, default=str)
            cursor.execute("UPDATE Blogs SET blog_comments = %s, blog_comments_count = COALESCE(blog_comments_count,0) + 1 WHERE blog_url = %s", (comments_json, blog_url))
            conn.commit()

        return jsonify(new_comment), 201

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass

@app.route("/api/blogs/<string:blog_url>/comments/<int:comment_id>", methods=["PUT"])
@token_required
def edit_comment(current_user, blog_url, comment_id):
    """
    Allows a logged-in user to edit their own comment.
    Expects JSON: { "comment_text": "new text here" }
    """
    try:
        data = request.get_json() or {}
        new_text = (data.get("comment_text") or "").strip()
        if not new_text:
            return jsonify({"error": "comment_text is required"}), 400

        conn = get_mysql_connection()
        with conn.cursor() as cursor:
            cursor.execute("SELECT blog_comments FROM Blogs WHERE blog_url = %s LIMIT 1", (blog_url,))
            row = cursor.fetchone()
            comments = _load_json_field(row or {}, "blog_comments")

            if not comments:
                return jsonify({"error": "No comments found"}), 404

            updated = False
            for c in comments:
                if int(c.get("comment_id")) == int(comment_id):
                    if str(c.get("customer_id")) != str(current_user):
                        return jsonify({"error": "Unauthorized: cannot edit others' comments"}), 403
                    c["comment_text"] = new_text
                    c["edited_at"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                    updated = True
                    break

            if not updated:
                return jsonify({"error": "Comment not found"}), 404

            comments_json = json.dumps(comments, default=str)
            cursor.execute("UPDATE Blogs SET blog_comments = %s WHERE blog_url = %s", (comments_json, blog_url))
            conn.commit()

        return jsonify({"message": "Comment updated successfully"}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass

@app.route("/api/blogs/<string:blog_url>/comments/<int:comment_id>", methods=["GET", "DELETE"])
@token_required
def delete_comment(current_user, blog_url, comment_id):
    """
    Allows a logged-in user to delete their own comment.
    Also decreases blog_comments_count by 1.
    """
    try:
        conn = get_mysql_connection()
        with conn.cursor() as cursor:
            cursor.execute("SELECT blog_comments, blog_comments_count FROM Blogs WHERE blog_url = %s LIMIT 1", (blog_url,))
            row = cursor.fetchone()
            comments = _load_json_field(row or {}, "blog_comments")

            if not comments:
                return jsonify({"error": "No comments found"}), 404

            deleted = False
            new_comments = []
            for c in comments:
                if int(c.get("comment_id")) == int(comment_id):
                    if str(c.get("customer_id")) != str(current_user):
                        return jsonify({"error": "Unauthorized: cannot delete others' comments"}), 403
                    deleted = True
                    continue  # skip adding this comment (delete)
                new_comments.append(c)

            if not deleted:
                return jsonify({"error": "Comment not found"}), 404

            new_count = max(0, len(new_comments))
            comments_json = json.dumps(new_comments, default=str)

            cursor.execute(
                "UPDATE Blogs SET blog_comments = %s, blog_comments_count = %s WHERE blog_url = %s",
                (comments_json, new_count, blog_url)
            )
            conn.commit()

        return jsonify({"message": "Comment deleted successfully"}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass


@app.route("/api/blogs/<string:blog_url>/like", methods=["POST"])
@token_required
def toggle_like(current_user, blog_url):
    
    try:
        if not request.is_json:
            return jsonify({"error": "Request must be JSON"}), 400

        conn = get_mysql_connection()
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT blog_likes, blog_likes_count FROM Blogs WHERE blog_url = %s LIMIT 1",
                (blog_url,)
            )
            row = cursor.fetchone()
            likes = _load_json_field(row or {}, "blog_likes")

            # Initialize list if empty
            if not likes:
                likes = []

            # Normalize to integer user IDs
            likes_set = set(map(str, likes))  # store as strings for JSON consistency
            user_id = str(current_user)

            # Toggle like/unlike
            if user_id in likes_set:
                likes_set.remove(user_id)
                liked = False
            else:
                likes_set.add(user_id)
                liked = True

            new_likes = list(likes_set)
            new_count = len(new_likes)

            # Update Blogs.blog_likes JSON and count
            cursor.execute(
                "UPDATE Blogs SET blog_likes = %s, blog_likes_count = %s WHERE blog_url = %s",
                (json.dumps(new_likes), new_count, blog_url)
            )
            conn.commit()

        return jsonify({"liked": liked, "likes_count": new_count}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass


@app.route('/api/all_reminders', methods=['GET'])
@token_required
def get_all_reminders(customer_id):
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            query = """
                SELECT 
                    r.id, 
                    r.start_date, 
                    r.end_date, 
                    r.reminder_time, 
                    r.product_ids, 
                    r.customer_id,
                    GROUP_CONCAT(rs.taken_datetime ORDER BY rs.taken_datetime DESC) AS taken_history
                FROM Reminders r
                LEFT JOIN ReminderStatus rs ON r.id = rs.reminder_id
                WHERE r.customer_id = %s
                GROUP BY r.id
                ORDER BY r.id DESC
            """
            cursor.execute(query, (customer_id,))
            reminders = cursor.fetchall()

            # Serialize datetime and timedelta fields
            for reminder in reminders:
                reminder['start_date'] = str(reminder['start_date']) if reminder['start_date'] else None
                reminder['end_date'] = str(reminder['end_date']) if reminder['end_date'] else None
                reminder['reminder_time'] = str(reminder['reminder_time']) if reminder['reminder_time'] else None
                reminder['taken_history'] = (
                    [dt for dt in reminder['taken_history'].split(",")]
                    if reminder['taken_history'] else []
                )
        return jsonify(reminders), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()


@app.route('/api/mark_as_taken', methods=['POST'])
@token_required
def mark_as_taken(customer_id):
    data = request.get_json()
    reminder_id = data.get("reminder_id")

    if not reminder_id or not customer_id:
        return jsonify({"error": "reminder_id and customer_id are required"}), 400

    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            # Check if the reminder exists and belongs to the customer
            check_query = """
                SELECT id 
                FROM Reminders 
                WHERE id = %s AND customer_id = %s
            """
            cursor.execute(check_query, (reminder_id, customer_id))
            reminder = cursor.fetchone()

            if not reminder:
                return jsonify({"error": "Reminder not found or does not belong to the customer"}), 404

            # Insert the record into ReminderStatus
            insert_query = """
                INSERT INTO ReminderStatus (reminder_id, taken_datetime)
                VALUES (%s, NOW())
            """
            cursor.execute(insert_query, (reminder_id,))
            mysql.commit()

        return jsonify({"message": "Marked as taken successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()



@app.route('/api/check_reminder', methods=['POST'])
@token_required
def check_reminder_at_given_time(customer_id):
    data = request.get_json()
    given_time = data.get("given_time")  # Format: 'YYYY-MM-DD HH:MM:SS'

    if not given_time:
        return jsonify({"error": "given_time is required"}), 400

    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            query = """
                SELECT 
                    r.id, 
                    r.start_date, 
                    r.end_date, 
                    r.reminder_time, 
                    r.product_ids, 
                    r.customer_id, 
                    GROUP_CONCAT(rs.taken_datetime ORDER BY rs.taken_datetime DESC) AS taken_history,
                    c.customer_name AS customer_name
                FROM Reminders r
                LEFT JOIN ReminderStatus rs ON r.id = rs.reminder_id
                JOIN Customers c ON r.customer_id = c.customer_id
                WHERE r.customer_id = %s
                AND r.start_date <= %s 
                AND r.end_date >= %s 
                GROUP BY r.id
            """
            cursor.execute(query, (customer_id, given_time, given_time))
            reminders = cursor.fetchall()

            active_reminders = []
            missed_reminders = []
            
            # Convert given_time string to datetime object
            given_datetime = datetime.strptime(given_time, '%Y-%m-%d %H:%M:%S')
            given_time_only = given_datetime.time()  # Extract the time part from the given datetime

            for reminder in reminders:
                reminder['start_date'] = str(reminder['start_date']) if reminder['start_date'] else None
                reminder['end_date'] = str(reminder['end_date']) if reminder['end_date'] else None
                reminder['reminder_time'] = str(reminder['reminder_time']) if reminder['reminder_time'] else None
                reminder['taken_history'] = (
                    [dt for dt in reminder['taken_history'].split(",")]
                    if reminder['taken_history'] else []
                )

                # Check if the given_datetime's date is already in taken_history
                taken_dates = {datetime.strptime(dt, "%Y-%m-%d %H:%M:%S").date() for dt in reminder['taken_history']}
                if given_datetime.date() in taken_dates:
                    continue  # Skip if the given date is already in taken_history

                reminder_time = datetime.strptime(str(reminder['reminder_time']), "%H:%M:%S").time()
                
                # Combine reminder_time with a placeholder date to compare times
                reminder_datetime = datetime.combine(datetime.today(), reminder_time)

                # Calculate the time difference to see if it's within the 15-minute buffer
                time_diff = abs((given_datetime - reminder_datetime).total_seconds())

                if time_diff <= 900:  # Within the 15-minute buffer
                    active_reminders.append(reminder)
                elif reminder_datetime < given_datetime:  # Past reminder (comparison with full datetime)
                    missed_reminders.append(reminder)

            return jsonify({
                "active_reminders": active_reminders,
                "missed_reminders": missed_reminders,
            }), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()



@app.route('/api/delete_reminder', methods=['POST'])
@token_required
def delete_reminder(customer_id):
    data = request.get_json()
    reminder_id = data.get("reminder_id")

    if not reminder_id:
        return jsonify({"error": "reminder_id is required"}), 400

    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            delete_query = "DELETE FROM Reminders WHERE id = %s AND customer_id = %s"
            cursor.execute(delete_query, (reminder_id, customer_id))
            mysql.commit()
        return jsonify({"message": "Reminder deleted successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()


@app.route('/api/add_reminder', methods=['POST'])
@token_required
def add_reminder(customer_id):
    data = request.get_json()
    required_fields = ['start_date', 'end_date', 'time', 'products']
    
    if not all(field in data for field in required_fields):
        return jsonify({"error": "Missing required fields"}), 400

    # Ensure that 'products' is a valid JSON format
    try:
        products = json.dumps(data['products'])  # Ensure products is serialized as a valid JSON string
    except TypeError:
        return jsonify({"error": "Invalid products format"}), 400

    try:
        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            rem_time = datetime.strptime(data['time'], "%Y-%m-%dT%H:%M:%S.%fZ").replace(second=0, microsecond=0).time()
            query = """
                INSERT INTO Reminders (start_date, end_date, reminder_time, product_ids, customer_id)
                VALUES (%s, %s, %s, %s, %s)
            """
            cursor.execute(query, (
                data['start_date'],
                data['end_date'],
                rem_time,
                products,  # Pass the serialized JSON
                customer_id
            ))
            mysql.commit()
            reminder_id = cursor.lastrowid

            cursor.execute("SELECT endpoint, p256dh, auth FROM PushSubscriptions WHERE user_id = %s", (customer_id,))
            subscriptions = cursor.fetchall()

            for subscription in subscriptions:
                message = "Hi, gentle reminder to take your medicine. Taking it now helps keep you on your path to wellness!"
                start_date = datetime.strptime(data['start_date'], "%Y-%m-%dT%H:%M:%S.%fZ").date()
                end_date = datetime.strptime(data['end_date'], "%Y-%m-%dT%H:%M:%S.%fZ").date()
                
                date_list = [(start_date + timedelta(days=i)).strftime("%Y-%m-%d") 
                                for i in range((end_date - start_date).days + 1)]

                for in_date in date_list:    

                    api_data = {
                        "payload": {
                            "title": "Medingen",
                            "body": message,
                            "icon": "/android-chrome-192x192.png",
                            "target_url": "https://medingen.in/reminder",
                            "reminder_id": reminder_id
                        },
                        "subscription": subscription,
                        "schedule_time": in_date + " " + rem_time.strftime("%H:%M:%S")
                    }
                    
                    api_url = f"http://ec2-35-154-224-159.ap-south-1.compute.amazonaws.com/send_notification_247832438"
                    response = requests.post(api_url, json=api_data)
                    print(f"Notification sent, status code: {response.status_code}")


        return jsonify({"message": "Reminder added successfully"}), 200
    except Exception as e:
        raise e
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()



@app.route('/api/notifications/<int:notification_id>/mark-as-read', methods=['PUT'])
def mark_notification_as_read(notification_id):
    try:
        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            update_sql = """
            UPDATE Notifications
            SET read_status = %s
            WHERE notification_id = %s
            """
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cursor.execute(update_sql, (current_time, notification_id))
            mysql.commit()

            return jsonify({"message": f"Notification {notification_id} marked as read."}), 200
    except Exception as e:
        return jsonify({"message": str(e)}), 500
    finally:
        cursor.close()




@app.route('/api/save-subscription', methods=['POST'])
@token_required
def save_subscription(user_id):
    try:
        data = request.get_json()
        data = data["subscription"]
        # Validate the required fields
        required_fields = ['endpoint', 'keys']
        if not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields"}), 400

        # Ensure `keys` contains `p256dh` and `auth`
        if not all(key in data['keys'] for key in ['p256dh', 'auth']):
            return jsonify({"error": "Missing required keys"}), 400

        # Extract subscription details
        endpoint = data['endpoint']
        p256dh = data['keys']['p256dh']
        auth = data['keys']['auth']

        # Save the subscription in the database
        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            query = """
                INSERT INTO PushSubscriptions (user_id, endpoint, p256dh, auth)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE endpoint = VALUES(endpoint), p256dh = VALUES(p256dh), auth = VALUES(auth)
            """
            cursor.execute(query, (user_id, endpoint, p256dh, auth))
            mysql.commit()

        return jsonify({"message": "Subscription saved successfully"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if 'mysql' in locals():
            mysql.close()


BUCKET = "medingen-store-new"
KEY = "redirects/redirect_map_new.json"
redirect_map = None

def load_redirect_map():
    global redirect_map
    if redirect_map is None:
        s3 = boto3.client("s3", region_name="ap-south-1")
        obj = s3.get_object(Bucket=BUCKET, Key=KEY)
        redirect_map = json.loads(obj["Body"].read())
    return redirect_map


@app.route('/blog/<path:path>', methods=['GET'])
def old_blog_handler(path):
    
    redirects = load_redirect_map()
    full_path = f"https://medingen.in/blog/{path}"
    if full_path in redirects:
        return redirect(redirects[full_path], code=302)
    else:
        return redirect("https://medingen.in/blogs/", code=302)

# Flask route for handling the /blogs/<path> URL
@app.route('/blogs/<path:path>', methods=['GET'])
def blog_handler(path):
    s3 = boto3.client('s3', region_name='ap-south-1' )
    
    redirects = load_redirect_map()
    full_path = f"https://medingen.in/blogs/{path}"
    if full_path in redirects:
        return redirect(redirects[full_path], code=302)

    # Fetch data from MySQL
    try:
        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            query = """
                SELECT id, blog_name, blog_image_url, blog_category_id, 
                       blog_description_url, blog_visit_count, blog_rc_priority, 
                       blog_created_date, meta_keywords, blog_url, meta_description, meta_title, 
                       meta_author, meta_author_title, meta_author_profile_url, blog_updated_date
                FROM Blogs
                WHERE blog_url = %s
                ORDER BY id DESC
            """
            cursor.execute(query, (path,))
            blogs = cursor.fetchall()

        # If no blog is found, set default values
        if not blogs:
            blog_data = {
                'blog_name': "Medingen Blogs",
                'meta_title': "Medingen Blog",
                'meta_description': "Medingen Online E-Commerce Platform for Generic medicines - Saves your health and wealth.",
                'meta_keywords': "",
                'blog_url': path,
                'blog_image_url': "medingen-logo.jpg",
                'meta_author': "Medingen Team",
                'meta_author_title': "",
                'meta_author_profile_url': "",
                'blog_created_date': "",
                'blog_updated_date': ""
            }
        else:
            blog_data = blogs[0] 
            blog_data = {
                'blog_name': blog_data['blog_name'],
                "meta_title": blog_data['meta_title'] or blog_data['blog_name'],
                'meta_description': blog_data['meta_description'] or "Medingen Online E-Commerce Platform for Generic medicines - Saves your health and wealth.",
                'meta_keywords': blog_data['meta_keywords'] or "",
                'blog_url': blog_data['blog_url'],
                'blog_image_url': blog_data['blog_image_url'] or "medingen-logo.jpg",
                'meta_author': blog_data['meta_author'] or "Medingen Team",
                'meta_author_title': blog_data['meta_author_title'] or "",
                'meta_author_profile_url': blog_data['meta_author_profile_url'] or "",
                'blog_created_date': blog_data['blog_created_date'] or "",
                'blog_updated_date': blog_data['blog_updated_date'] or ""
            }

        # Fetch the index.html from S3
        response = s3.get_object(Bucket='medingen-in-new-2025', Key='index.html')
        html = response['Body'].read().decode('utf-8')

        def update_or_add_meta(html, pattern, replacement, meta_tag):
            if re.search(pattern, html):
                return re.sub(pattern, replacement, html)
            else:
                head_close = '</head>'
                return html.replace(head_close, f"\n    {meta_tag}\n{head_close}")

        # Modify meta tags and title in the HTML
        modified_html = update_or_add_meta(
            html,
            r'<meta name="description" content=".*?">',
            f'<meta name="description" content="{blog_data["meta_description"]}">',
            f'<meta name="description" content="{blog_data["meta_description"]}">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta name="keywords" content=".*?">',
            f'<meta name="keywords" content="{blog_data["meta_keywords"]}">',
            f'<meta name="keywords" content="{blog_data["meta_keywords"]}">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta name="author" content=".*?">',
            f'<meta name="author" content="{blog_data["meta_author"]}">',
            f'<meta name="author" content="{blog_data["meta_author"]}">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta name="robots" content=".*?">',
            '<meta name="robots" content="index, follow">',
            '<meta name="robots" content="index, follow">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta property="og:title" content=".*?">',
            f'<meta property="og:title" content="{blog_data["blog_name"]}">',
            f'<meta property="og:title" content="{blog_data["blog_name"]}">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta property="og:description" content=".*?">',
            f'<meta property="og:description" content="{blog_data["meta_description"]}">',
            f'<meta property="og:description" content="{blog_data["meta_description"]}">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta property="og:title" content=".*?">',
            f'<meta property="og:title" content="{blog_data["meta_title"]}">',
            f'<meta property="og:title" content="{blog_data["meta_title"]}">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta property="og:url" content=".*?">',
            f'<meta property="og:url" content="https://medingen.in/blogs/{blog_data["blog_url"]}">',
            f'<meta property="og:url" content="https://medingen.in/blogs/{blog_data["blog_url"]}">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta property="og:type" content=".*?">',
            '<meta property="og:type" content="article">',
            '<meta property="og:type" content="article">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta property="og:image" content=".*?">',
            f'<meta property="og:image" content="https://d1dh0rr5xj2p49.cloudfront.net/blogs/images/{blog_data["blog_image_url"]}">',
            f'<meta property="og:image" content="https://d1dh0rr5xj2p49.cloudfront.net/blogs/images/{blog_data["blog_image_url"]}">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta name="twitter:card" content=".*?">',
            f'<meta name="twitter:card" content="https://d1dh0rr5xj2p49.cloudfront.net/blogs/images/{blog_data["blog_image_url"]}">',
            f'<meta name="twitter:card" content="https://d1dh0rr5xj2p49.cloudfront.net/blogs/images/{blog_data["blog_image_url"]}">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta name="twitter:title" content=".*?">',
            f'<meta name="twitter:title" content="{blog_data["blog_name"]}">',
            f'<meta name="twitter:title" content="{blog_data["blog_name"]}">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta name="twitter:description" content=".*?">',
            f'<meta name="twitter:description" content="{blog_data["meta_description"]}">',
            f'<meta name="twitter:description" content="{blog_data["meta_description"]}">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<meta name="twitter:image" content=".*?">',
            f'<meta name="twitter:image" content="https://d1dh0rr5xj2p49.cloudfront.net/blogs/images/{blog_data["blog_image_url"]}">',
            f'<meta name="twitter:image" content="https://d1dh0rr5xj2p49.cloudfront.net/blogs/images/{blog_data["blog_image_url"]}">'
        )

        modified_html = update_or_add_meta(
            modified_html,
            r'<link rel="canonical" href=".*?">',
            f'<link rel="canonical" href="https://medingen.in/blogs/{blog_data["blog_url"]}">',
            f'<link rel="canonical" href="https://medingen.in/blogs/{blog_data["blog_url"]}">'
        )

        modified_html = modified_html + f'''
            <script type="application/ld+json">
            {{
                "@context": "https://schema.org",
                "@type": "Article",
                "headline": "{blog_data["meta_title"]}",
                "description": "{blog_data["meta_description"]}",
                "image": "https://d1dh0rr5xj2p49.cloudfront.net/blogs/images/{blog_data["blog_image_url"]}",
                "author": {{
                    "@type": "Person",
                    "name": "{blog_data["meta_author"]}",
                    "jobTitle": "{blog_data["meta_author_title"]}",
                    "url": "{blog_data["meta_author_profile_url"]}"
                }},
                "publisher": {{
                    "@type": "Organization",
                    "name": "Medingen",
                    "logo": {{
                        "@type": "ImageObject",
                        "url": "https://medingen.in/migfulllogo.png"
                    }}
                }},
                "datePublished": "{blog_data["blog_created_date"]}",
                "dateModified": "{blog_data["blog_updated_date"]}",
                "mainEntityOfPage": {{
                    "@type": "WebPage",
                    "@id": "https://medingen.in/blogs/{blog_data["blog_url"]}"
                }}
            }}
            </script>
            '''

        modified_html = re.sub(
            r'<title>.*?</title>',
            f'<title>{blog_data["blog_name"]} - Medingen</title>',
            modified_html
        )

        return Response(modified_html, mimetype='text/html')

    except Exception as e:
        error_message = f"Error fetching content: {str(e)}"
        stack_trace = traceback.format_exc()  
        print(f"{error_message}\n{stack_trace}") 
        return Response(error_message, status=500)

from urllib.parse import quote

FEATURABLE_API_URL = "https://featurable.com/api/v1/widgets/fb210dba-0301-4000-982d-6e8006ca39f3"

@app.route('/product/<path:product_name>', methods=['GET'])
def product_handler(product_name):

    if "%" in product_name:
        encoded_slug = product_name.replace("%", "%25")
        return redirect(f"/product/{encoded_slug}", code=301)

    normalized_name = product_name.lower()
    if product_name != normalized_name:
        return redirect(f"/product/{normalized_name}", code=301)

    try:
        s3 = boto3.client('s3', region_name='ap-south-1')
        redirects = load_redirect_map()
        encoded_name = quote(normalized_name)
        full_path = f"https://medingen.in/product/{encoded_name}"

        if full_path in redirects and redirects[full_path] != "https://medingen.in/":
            return redirect(redirects[full_path], code=301)

        mysql = get_mysql_connection()
        with mysql.cursor() as cursor:
            query = """
                SELECT name AS product_name, product_id, 
                       photo, long_description,
                       meta_keywords, meta_description, meta_title, product_pricing_new as price,
                       manufacturer, composition, product_pricing_old as mrp, product_name_url,
                       salt_name, used_for, marketed_by, packaging, prescription_required
                FROM Products 
                WHERE LOWER(product_name_url) = LOWER(%s);
            """
            cursor.execute(query, (normalized_name,))
            product = cursor.fetchone()

        default_meta = {
            'product_name': "Medingen",
            'product_id': None,
            'photo': "[]",
            'long_description': None,
            'meta_title': "Buy Generic Medicines Online | Trusted Store - Medingen",
            'meta_description': "Medingen offers Generic Medicines Home Delivery at affordable prices. Order online from a trusted store and get fast doorstep delivery.",
            'meta_keywords': "",
            'product_url': f"https://medingen.in/product/{normalized_name}",
            'images': ["https://d1dh0rr5xj2p49.cloudfront.net/products/medingen-logo.jpg"],
            'price': 0,
            'manufacturer': "Unknown",
            'composition': "Unknown",
            'mrp': 0,
            'salt_name': "Unknown",
            'used_for': "Unknown",
            'marketed_by': "Unknown",
            'packaging': "Unknown",
            'prescription_required': "Unknown",
        }

        product_data = product if product else default_meta

        images = []
        try:
            if product and product.get("photo"):
                photo_data = json.loads(product["photo"])
                images = [
                    f"https://d1dh0rr5xj2p49.cloudfront.net/products/{img['img']}"
                    for img in photo_data if "img" in img
                ]
        except Exception as e:
            print(f"Image parsing error: {e}")

        if not images:
            images = default_meta['images']

        product_data['images'] = images
        product_data['image_url'] = images[0]
        product_data['product_url'] = f"https://medingen.in/product/{normalized_name}"

        def safe_decimal(val, default=0):
            if val is None:
                return default
            if isinstance(val, Decimal):
                return float(val)
            return val

        product_data['price'] = safe_decimal(product_data.get('price'), 0)
        product_data['mrp'] = safe_decimal(product_data.get('mrp'), 0)

        # ── Fetch live review count & rating from Featurable ──────────────────
        review_count = "52"       # fallback defaults
        average_rating = "5.0"
        try:
            review_resp = requests.get(FEATURABLE_API_URL, timeout=5)
            if review_resp.status_code == 200:
                review_data = review_resp.json()
                if review_data.get("success"):
                    review_count   = str(review_data.get("totalReviewCount", 52))
                    average_rating = str(float(review_data.get("averageRating", 5.0)))
        except Exception as e:
            print(f"Featurable API error: {e}")
        # ─────────────────────────────────────────────────────────────────────

        response = s3.get_object(Bucket='medingen-in-new-2025', Key='index.html')
        html = response['Body'].read().decode('utf-8')

        def update_or_add_meta(html, pattern, replacement, meta_tag):
            if re.search(pattern, html):
                return re.sub(pattern, replacement, html)
            else:
                head_close = '</head>'
                return html.replace(head_close, f"\n    {meta_tag}\n{head_close}")

        modified_html = update_or_add_meta(
            html, r'<meta name="description" content=".*?">',
            f'<meta name="description" content="{product_data["meta_description"]}">',
            f'<meta name="description" content="{product_data["meta_description"]}">'
        )
        modified_html = update_or_add_meta(
            modified_html, r'<meta name="keywords" content=".*?">',
            f'<meta name="keywords" content="{product_data["meta_keywords"]}">',
            f'<meta name="keywords" content="{product_data["meta_keywords"]}">'
        )
        modified_html = re.sub(
            r'<title>.*?</title>',
            f'<title>{product_data["meta_title"]}</title>',
            modified_html
        )

        canonical_url = f"https://medingen.in/product/{normalized_name}"
        modified_html = update_or_add_meta(
            modified_html,
            r'<link rel="canonical" href=".*?">',
            f'<link rel="canonical" href="{canonical_url}">',
            f'<link rel="canonical" href="{canonical_url}">'
        )

        fallback_data = {
            "indication": None,
            "sideEffects": [],
            "safety": None
        }

        try:
            if product and product.get("long_description"):
                long_desc_key = product["long_description"]
                long_description_url = f"https://d1dh0rr5xj2p49.cloudfront.net/product_description/{long_desc_key}"
                resp = requests.get(long_description_url, timeout=5)

                if resp.status_code == 200:
                    html_content = resp.text
                    clean_text = re.sub(r'<.*?>', ' ', html_content).strip()

                    if not product_data.get("used_for"):
                        match = re.search(r'(used for|treats|helps with)(.*?)(\.|\n)', clean_text, re.I)
                        if match:
                            fallback_data["indication"] = match.group(2).strip()

                    se_matches = re.findall(r'(side effects?.*?\.)', clean_text, re.I)
                    fallback_data["sideEffects"] = se_matches[:5]

                    warn_match = re.search(r'(warning|precaution)(.*?)(\.|\n)', clean_text, re.I)
                    if warn_match:
                        fallback_data["safety"] = warn_match.group(2).strip()

        except Exception as e:
            print("Fallback parsing error:", e)

        ld_json_data = {
            "@context": "https://schema.org",
            "@graph": [
                # 1. ORGANIZATION
                {
                    "@type": "Organization",
                    "@id": "https://medingen.in/#organization",
                    "name": "Medingen",
                    "url": "https://medingen.in/",
                    "logo": {
                        "@type": "ImageObject",
                        "url": "https://medingen.in/migfulllogo.svg"
                    },
                    "description": "Medingen is a tech-enabled platform to discover, compare, and save on medicines online.",
                    "address": {
                        "@type": "PostalAddress",
                        "streetAddress": "School street, Mangadu",
                        "addressLocality": "Chennai",
                        "addressRegion": "Tamil Nadu",
                        "postalCode": "600122",
                        "addressCountry": "IN"
                    },
                    "contactPoint": {
                        "@type": "ContactPoint",
                        "telephone": "+91-7090123709",
                        "contactType": "customer support"
                    },
                    "sameAs": [
                        "https://www.facebook.com/people/Medingen/61567679517972/",
                        "https://www.instagram.com/medin.gen/",
                        "https://www.youtube.com/@ashash_mig/"
                    ],
                    "aggregateRating": {
                        "@type": "AggregateRating",
                        "ratingValue": average_rating,
                        "reviewCount": review_count
                    }
                },

                # 2. WEBSITE
                {
                    "@type": "WebSite",
                    "@id": "https://medingen.in/#website",
                    "url": "https://medingen.in/",
                    "name": "Medingen",
                    "publisher": {
                        "@id": "https://medingen.in/#organization"
                    },
                    "potentialAction": {
                        "@type": "SearchAction",
                        "target": "https://medingen.in/search?q={search_term_string}",
                        "query-input": "required name=search_term_string"
                    }
                },

                # 3. WEBPAGE (dynamic)
                {
                    "@type": "WebPage",
                    "@id": f"{product_data['product_url']}#webpage",
                    "url": product_data['product_url'],
                    "name": product_data['meta_title'],
                    "description": product_data['meta_description'],
                    "inLanguage": "en-IN",
                    "isPartOf": {
                        "@id": "https://medingen.in/#website"
                    },
                    "mainEntity": {
                        "@id": f"{product_data['product_url']}#product"
                    },
                    "publisher": {
                        "@id": "https://medingen.in/#organization"
                    }
                },

                # 4. PRODUCT
                {
                    "@type": "Product",
                    "@id": f"{product_data['product_url']}#product",
                    "name": product_data['product_name'],
                    "url": product_data['product_url'],
                    "description": product_data['meta_description'],
                    "image": images,
                    "sku": product_data['product_id'] or "SKU-UNKNOWN",
                    "category": "Medicines",
                    "brand": {
                        "@type": "Brand",
                        "name": product_data['manufacturer'] or "Unknown"
                    },
                    "manufacturer": {
                        "@type": "Organization",
                        "name": product_data['manufacturer'] or "Unknown"
                    },
                    "activeIngredient": product_data['salt_name'],
                    "dosageForm": product_data.get('consume_type') or "Tablet",
                    "indication": product_data['used_for'] or fallback_data["indication"] or "General pain relief",
                    "administrationRoute": "Oral",
                    "additionalProperty": [
                        {
                            "@type": "PropertyValue",
                            "name": "Salt",
                            "value": product_data['salt_name']
                        },
                        {
                            "@type": "PropertyValue",
                            "name": "Pack Size",
                            "value": product_data['packaging']
                        },
                        {
                            "@type": "PropertyValue",
                            "name": "Prescription Required",
                            "value": product_data['prescription_required']
                        },
                        {
                            "@type": "PropertyValue",
                            "name": "Marketed By",
                            "value": product_data['marketed_by']
                        }
                    ],
                    "aggregateRating": {
                        "@type": "AggregateRating",
                        "ratingValue": average_rating,
                        "reviewCount": review_count
                    },
                    "offers": {
                        "@type": "Offer",
                        "url": product_data['product_url'],
                        "priceCurrency": "INR",
                        "price": float(product_data.get('price', 0)),
                        "availability": "https://schema.org/InStock" if product_data['price'] > 0 else "https://schema.org/OutOfStock",
                        "itemCondition": "https://schema.org/NewCondition",
                        "seller": {
                            "@id": "https://medingen.in/#organization"
                        }
                    }
                }
            ]
        }

        if fallback_data["safety"]:
            ld_json_data["warning"] = fallback_data["safety"]

        product_html = f"""
        <section class="product-detail">
            <h1>{product_data['product_name']}</h1>
            <img src="{product_data['image_url']}" alt="{product_data['product_name']}" style="max-width:300px">
            <p><strong>Price:</strong> ₹{product_data['price']}</p>
            <p><strong>MRP:</strong> ₹{product_data['mrp']}</p>
            <p><strong>Manufacturer:</strong> {product_data['manufacturer']}</p>
            <p><strong>Composition:</strong> {product_data['composition']}</p>
            <p><strong>Salt:</strong> {product_data['salt_name']}</p>
            <p><strong>Used For:</strong> {product_data['used_for']}</p>
            <p><strong>Packaging:</strong> {product_data['packaging']}</p>
            <p><strong>Marketed By:</strong> {product_data['marketed_by']}</p>
            <p><strong>Prescription Required:</strong> {product_data['prescription_required']}</p>
            <p>{product_data['meta_description']}</p>
        </section>
        """

        modified_html = modified_html.replace(
            "</body>",
            product_html +
            f'<script type="application/ld+json">{json.dumps(ld_json_data)}</script>\n</body>'
        )

        try:
            if product and product.get("long_description"):
                long_desc_key = product["long_description"]
                long_description_url = f"https://d1dh0rr5xj2p49.cloudfront.net/product_description/{long_desc_key}"
                resp = requests.get(long_description_url, timeout=5)
                if resp.status_code == 200:
                    long_desc_html = resp.text
                    faq_matches = []
                    faq_matches += re.findall(r'<p><strong>Q\.\s*(.*?)</strong></p>\s*<p>(.*?)</p>', long_desc_html, flags=re.S)
                    faq_matches += re.findall(r'<p><strong>Q\.\s*(.*?)</strong><br>(.*?)</p>', long_desc_html, flags=re.S)

                    main_entities = []
                    for question, answer in faq_matches:
                        q = re.sub(r'<.*?>', '', question.strip())
                        a = re.sub(r'<.*?>', '', answer.strip())
                        main_entities.append({
                            "@type": "Question",
                            "name": q,
                            "acceptedAnswer": {"@type": "Answer", "text": a}
                        })

                    if main_entities:
                        faq_jsonld = {
                            "@context": "https://schema.org",
                            "@type": "FAQPage",
                            "mainEntity": main_entities[:7]
                        }
                        modified_html = modified_html.replace(
                            "</body>",
                            f'<script type="application/ld+json">{json.dumps(faq_jsonld)}</script>\n</body>'
                        )
        except Exception as e:
            print(f"FAQ parsing error: {e}")

        return Response(modified_html, mimetype='text/html')

    except Exception as e:
        error_message = f"Error fetching content: {str(e)}"
        stack_trace = traceback.format_exc()
        return Response(error_message, status=500)


@app.route("/api/footer-products", methods=['GET'])
def get_footer_products():
    try:
        connection = get_mysql_connection()

        products_map = {
            "topSellingMedicines": [],
            "topHealthcareDevices": [],
            "topHealthProducts": [],
            "popularMedicine": [],
            "frequentlySearchedMedicine": [],
            "dealsForTheDay": []
        }

        with connection.cursor() as cursor:

            # ✅ LIMIT PER CATEGORY (MySQL 8+)
            sql = """
                SELECT * FROM (
                    SELECT 
                        f.category,
                        p.product_id,
                        p.name,
                        p.product_name_url,

                        JSON_UNQUOTE(JSON_EXTRACT(p.photo, '$[0].img')) AS first_image_url,
                        p.product_pricing_new,
                        p.product_pricing_old,
                        p.salt_name,
                        p.composition,
                        p.categories,
                        p.inStock,
                        p.rc,
                        p.manufacturer,

                        ROW_NUMBER() OVER (PARTITION BY f.category ORDER BY f.id DESC) as rn

                    FROM FooterDisplayItems f
                    JOIN Products p ON f.product_name = p.name

                    WHERE p.product_name_url IS NOT NULL 
                      AND p.product_name_url != ''
                      AND p.product_pricing_new IS NOT NULL
                      AND p.visibility_status = 'Published'

                ) ranked
                WHERE rn <= 20
            """

            cursor.execute(sql)
            results = cursor.fetchall()

            for row in results:
                item = {
                    # ✅ existing fields (UNCHANGED)
                    "label": row["name"],
                    "type": "internal",
                    "value": f"/product/{row['product_name_url']}",

                    # ✅ extra fields (UNCHANGED STRUCTURE)
                    "product_id": row["product_id"],
                    "first_image_url": row["first_image_url"],
                    "product_pricing_new": row["product_pricing_new"],
                    "product_pricing_old": row["product_pricing_old"],
                    "salt_name": row["salt_name"],
                    "composition": row["composition"],
                    "categories": row["categories"],
                    "inStock": row["inStock"],
                    "rc": row["rc"],
                    "manufacturer": row["manufacturer"],
                    "product_name_url": row["product_name_url"]
                }

                cat = row["category"]

                if cat in products_map:
                    products_map[cat].append(item)

        connection.close()

        return jsonify(products_map), 200

    except Exception as e:
        print(f"Error fetching footer products: {e}")
        return jsonify({"error": str(e)}), 500

        
@app.route('/api/category_hierarchy', methods=['GET'])
def get_category_hierarchy():
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            cursor.execute(
                "SELECT hierarchy_data FROM CategoryHierarchy ORDER BY id DESC LIMIT 1"
            )
            result = cursor.fetchone()

            if result and result['hierarchy_data']:
                return jsonify(json.loads(result['hierarchy_data'])), 200
            else:
                return jsonify([]), 200

    except Exception as e:
        print("Error fetching hierarchy:", e)
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()


@app.route('/api/main_categories', methods=['GET'])
def get_main_categories():
    mysql = get_mysql_connection()
    try:
        with mysql.cursor() as cursor:
            sql = """
                SELECT 
                    mc.id AS main_id, 
                    mc.name AS main_title, 
                    c.category_name AS sub_name, 
                    c.category_image_url AS sub_image
                FROM MainCategory mc
                LEFT JOIN Category c ON c.main_category = mc.name
                ORDER BY mc.name ASC, c.category_name ASC
            """
            cursor.execute(sql)
            rows = cursor.fetchall()

            grouped_categories = {}

            def normalize_key(name):
                if not name:
                    return None
                return (
                    name.strip()
                        .lower()
                        .replace('&', 'and')
                        .replace('  ', ' ')
                )

            def normalize_display(name):
                if not name:
                    return None

                n = normalize_key(name)

                mapping = {
                    ('infectious care', 'infective care'): 'Infectious Care',
                    ('cardiac', 'heart care'): 'Heart Care',
                    ('oncology care', 'cancer care'): 'Cancer Care',
                    ('urology care', 'urological care'): 'Urology Care',
                    ('digestive care',): 'Digestive Care',
                    ('hair care',): 'Hair Care',
                    ('foot care',): 'Foot Care',
                    ('uti care',): 'UTI Care',
                    ('respiratory care',): 'Respiratory Care',
                    ('eye care',): 'Eye Care',
                    ('cold and cough', 'cold and cough care'): 'Cold & Cough'
                }

                for keys, value in mapping.items():
                    if n in keys:
                        return value

                return name.strip().title()

            for row in rows:
                main_id = row['main_id']

                if main_id not in grouped_categories:
                    grouped_categories[main_id] = {
                        "id": main_id,
                        "name": row['main_title'],
                        "sub_categories": [],
                        "_seen": {}
                    }

                raw_name = row['sub_name']
                key = normalize_key(raw_name)
                display_name = normalize_display(raw_name)
                image = row['sub_image']

                if not key:
                    continue

                seen = grouped_categories[main_id]["_seen"]

                if key not in seen:
                    seen[key] = {
                        "name": display_name,
                        "image": image or ""
                    }
                else:
                    if not seen[key]["image"] and image:
                        seen[key]["image"] = image

            result = []
            for cat in grouped_categories.values():
                cat["sub_categories"] = list(cat["_seen"].values())
                del cat["_seen"]
                result.append(cat)

            return jsonify(result), 200

    except Exception as e:
        print("Error fetching main categories:", e)
        return jsonify({"error": str(e)}), 500
    finally:
        mysql.close()




@app.errorhandler(404)
def resource_not_found(e):
    return make_response(jsonify(Message='Hello medingen...!'), 404)


if __name__ == '__main__':
    app.run(host="0.0.0.0", port="8001", debug=True)
